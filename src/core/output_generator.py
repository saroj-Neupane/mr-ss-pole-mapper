import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from .utils import Utils


class OutputGenerator:
    """Handles Excel output generation"""
    
    def __init__(self, config, mapping_data=None, qc_reader=None):
        self.config = config
        self.mapping_data = mapping_data or []
        self.qc_reader = qc_reader
        self._processed_data_cache = {}
    
    def write_output(self, result_data, output_file):
        """Write processed data to Excel output file"""
        try:
            # Filter out empty or invalid data
            filtered_data = [row for row in result_data if row and row.get('Pole')]
            
            if not filtered_data:
                logging.warning("No valid data to write after filtering")
                return

            # Sort data using shared utility function
            sorted_data = sorted(filtered_data, key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))

            # Create data cache for QC sheet population
            self._processed_data_cache = {}
            for row in sorted_data:
                pole = row.get('Pole', '').strip()
                if pole:
                    self._processed_data_cache[pole] = row

            # Validate the output file before loading
            output_path = Path(output_file)
            if not output_path.exists() or output_path.stat().st_size == 0:
                logging.error(f"Output file '{output_file}' is missing or empty.")
                return

            # Attempt to load the workbook inside a try/except block to catch EOFError
            try:
                # Use keep_vba=True to preserve macros in .xlsm files
                wb = load_workbook(output_file, keep_vba=True)
            except EOFError as eof_error:
                logging.error(f"EOFError encountered when loading workbook '{output_file}': {eof_error}. The template file may be corrupted.")
                return

            # Determine worksheet to use
            if hasattr(self, 'config') and self.config:
                worksheet_name = self.config.get('output_settings', {}).get('worksheet_name', 'Consumers pg1')
            else:
                worksheet_name = 'Consumers pg1'

            if worksheet_name in wb.sheetnames:
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                logging.warning(f"Worksheet '{worksheet_name}' not found, using '{ws.title}'")

            # Write data; using mapped writing if available, else a simple write
            if hasattr(self, 'mapping_data') and self.mapping_data:
                self._write_data_to_worksheet(ws, sorted_data, self.mapping_data)
            else:
                self._write_data_simple(ws, sorted_data)

            # Automatically populate QC sheet if QC reader is active
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("QC reader is active - populating QC sheet")
                self._populate_qc_sheet(wb)
                
                # Add conditional formatting to compare main sheet and QC sheet
                logging.info("Adding conditional formatting to compare main sheet and QC sheet")
                self._add_sheet_comparison_formatting(wb, worksheet_name)
            else:
                logging.info("QC reader not active - skipping QC sheet population")

            wb.save(output_file)
            logging.info(f"Successfully wrote {len(sorted_data)} records to {output_file}")

        except Exception as e:
            logging.error(f"Error writing output: {e}")
            raise
    
    def _write_data_to_worksheet(self, ws, sorted_data, mapping_data):
        """Write data to worksheet using column mappings"""
        # Get output settings
        output_settings = self.config.get('output_settings', {})
        header_row = output_settings.get('header_row', 2)
        data_start_row = output_settings.get('data_start_row', 3)
        
        # Create mapping from output column names to Excel column indices
        column_mapping = {}
        for element, attribute, output_column in mapping_data:
            if output_column not in column_mapping:
                # Find the column in the header row
                for col_idx in range(1, ws.max_column + 1):
                    header_cell = ws.cell(row=header_row, column=col_idx)
                    if header_cell.value and str(header_cell.value).strip() == output_column:
                        column_mapping[output_column] = col_idx
                        break
        
        # Write data rows
        for row_idx, row_data in enumerate(sorted_data):
            excel_row = data_start_row + row_idx
            
            # Set line number
            if 'Line No.' in column_mapping:
                ws.cell(row=excel_row, column=column_mapping['Line No.']).value = row_idx + 1
            
            # Write mapped data
            for element, attribute, output_column in mapping_data:
                if output_column in column_mapping:
                    col_idx = column_mapping[output_column]
                    
                    # Get the internal key for this mapping
                    internal_key = self._get_internal_key(element, attribute)
                    value = row_data.get(internal_key, "")
                    
                    ws.cell(row=excel_row, column=col_idx).value = value
        
        logging.info(f"Wrote {len(sorted_data)} rows using column mappings")
    
    def _write_data_simple(self, ws, sorted_data):
        """Simple data writing without mappings"""
        # Find the first row with data to determine column structure
        if not sorted_data:
            return
        
        # Use the keys from the first row as column headers
        headers = list(sorted_data[0].keys())
        
        # Write headers (assuming row 2)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_idx).value = header
        
        # Write data (starting from row 3)
        for row_idx, row_data in enumerate(sorted_data):
            excel_row = 3 + row_idx
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=excel_row, column=col_idx).value = value
        
        logging.info(f"Wrote {len(sorted_data)} rows using simple format")
    
    def _get_internal_key(self, element, attribute):
        """Get the internal key used in row data for a given element/attribute mapping"""
        # Handle special cases
        if element == "Pole" and attribute == "SCID":
            return "Pole"
        elif element == "Pole" and attribute == "To Pole":
            return "To Pole"
        elif element == "Pole" and attribute == "Line No.":
            return "Line No."
        elif element == "Pole" and attribute == "Span Distance":
            return "Span Length"
        elif element == "Pole" and attribute == "Pole Height/Class":
            return "Pole Height & Class"
        elif element == "Pole" and attribute == "Address":
            return "Pole Address"
        elif element == "Pole" and attribute == "Guy Info":
            return "Guy Direction"  # Or could be "Guy Lead" depending on mapping
        elif element == "Pole" and attribute == "Existing Risers":
            return "Existing Risers"
        elif element == "Power" and attribute == "Height":
            return "Power Height"
        elif element == "Power" and attribute == "Midspan":
            return "Power Midspan"
        elif element == "Streetlight" and attribute == "Height":
            return "Streetlight (bottom of bracket)"
        elif element == "Street Light" and attribute == "Height":
            return "Street Light Height"
        elif element in ["comm1", "comm2", "comm3", "comm4"] and attribute == "Height":
            return element
        elif element == "All_Comm_Heights" and attribute == "Summary":
            return "All Communication Heights"
        elif element == "Total_Comm_Count" and attribute == "Count":
            return "Total Communication Count"
        elif element in self.config.get("telecom_providers", []) and attribute == "Attachment Ht":
            return element
        else:
            # Default: use element name as key
            return element
    
    def _populate_qc_sheet(self, workbook):
        """Populate QC sheet with data from QC file"""
        # This is a complex method that would be moved from PoleDataProcessor
        # For now, we'll keep a simplified version
        try:
            if "QC" not in workbook.sheetnames:
                logging.info("No existing QC sheet found, skipping QC data population")
                return
            
            logging.info("QC sheet population would be implemented here")
            # Implementation would be moved from PoleDataProcessor
            
        except Exception as e:
            logging.error(f"Error populating QC sheet: {e}")
    
    def _add_sheet_comparison_formatting(self, workbook, main_sheet_name):
        """Add conditional formatting to compare sheets"""
        try:
            logging.info("Sheet comparison formatting would be implemented here")
            # Implementation would be moved from PoleDataProcessor
            
        except Exception as e:
            logging.error(f"Error adding sheet comparison formatting: {e}")
    
    def generate_output_file(self, job_name, template_path):
        """Generate output file from template"""
        try:
            template_path = Path(template_path)
            if not template_path.exists():
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            # Create output filename
            output_filename = f"{job_name}_MR_SS.xlsx"
            output_path = template_path.parent / output_filename
            
            # Copy template to output location
            import shutil
            shutil.copy2(template_path, output_path)
            
            logging.info(f"Generated output file: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error generating output file: {e}")
            raise 