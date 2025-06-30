import pandas as pd
import logging
from pathlib import Path


class QCReader:
    """Reads and processes QC (Quality Control) Excel files for pole connection filtering"""
    
    def __init__(self, qc_file_path=None, ignore_scid_keywords=None):
        """
        Initialize QC reader with optional QC file path and ignore keywords
        
        Args:
            qc_file_path (str, optional): Path to QC Excel file
            ignore_scid_keywords (list, optional): Keywords to ignore when normalizing SCIDs
        """
        self.qc_file_path = qc_file_path
        self.ignore_scid_keywords = ignore_scid_keywords or []
        self.connections = set()  # Set of (from_scid, to_scid) tuples (normalized for matching)
        self.ordered_connections = []  # List preserving QC file order (normalized for matching)
        self.original_ordered_connections = []  # List preserving EXACT QC file format
        self.qc_data_rows = []  # Complete row data from QC file
        self.qc_scids = set()  # All SCIDs mentioned in QC file (normalized)
        self._active = False
        
        if qc_file_path:
            self.load_qc_file(qc_file_path)
    
    def load_qc_file(self, qc_file_path):
        """
        Load QC data from Excel file
        
        Args:
            qc_file_path (str): Path to QC Excel file
        """
        try:
            qc_path = Path(qc_file_path)
            if not qc_path.exists():
                logging.warning(f"QC file not found: {qc_file_path}")
                self._active = False
                return
            
            # Read QC Excel file - get all sheet names first
            import openpyxl
            wb = openpyxl.load_workbook(qc_file_path, data_only=True)
            sheet_names = wb.sheetnames
            logging.info(f"QC file has {len(sheet_names)} sheets: {sheet_names}")
            
            all_connections = []
            all_data_rows = []
            sheets_processed = 0
            
            # Process each sheet
            for sheet_name in sheet_names:
                try:
                    # Try different header row positions (common positions are 0, 1, 2)
                    df = None
                    for header_row in [2, 0, 1]:  # Try row 3 first (index 2), then row 1 (index 0), then row 2 (index 1)
                        try:
                            temp_df = pd.read_excel(qc_file_path, sheet_name=sheet_name, header=header_row, dtype=str, keep_default_na=False)
                            # Check if this header row has the required columns
                            if 'Pole' in temp_df.columns and 'To Pole' in temp_df.columns:
                                df = temp_df
                                logging.debug(f"Sheet '{sheet_name}': Found headers at row {header_row + 1}")
                                break
                        except:
                            continue
                    
                    if df is None:
                        # If no valid header row found, try reading without header and look for Pole/To Pole columns
                        df = pd.read_excel(qc_file_path, sheet_name=sheet_name, header=None, dtype=str, keep_default_na=False)
                        # Look for rows containing 'Pole' and 'To Pole'
                        for idx, row in df.iterrows():
                            if 'Pole' in row.values and 'To Pole' in row.values:
                                # Use this row as header
                                df.columns = row
                                df = df.iloc[idx+1:].reset_index(drop=True)
                                logging.debug(f"Sheet '{sheet_name}': Found headers at row {idx + 1}")
                                break
                        else:
                            logging.info(f"Sheet '{sheet_name}' missing required columns 'Pole' and 'To Pole' - skipping")
                            continue

                    
                    # Process connections from this sheet
                    sheet_connections = []
                    sheet_data_rows = []
                    for _, row in df.iterrows():
                        from_pole_orig = str(row['Pole']).strip()
                        to_pole_orig = str(row['To Pole']).strip()
                        
                        # Skip empty rows
                        if not from_pole_orig or not to_pole_orig or from_pole_orig == 'nan' or to_pole_orig == 'nan':
                            continue
                        
                        sheet_connections.append((from_pole_orig, to_pole_orig))
                        
                        # Store complete row data as dictionary
                        row_data = {}
                        for col in df.columns:
                            value = str(row[col]).strip() if pd.notna(row[col]) and str(row[col]).strip() != 'nan' else ''
                            row_data[col] = value
                        sheet_data_rows.append(row_data)
                    
                    all_connections.extend(sheet_connections)
                    all_data_rows.extend(sheet_data_rows)
                    sheets_processed += 1
                    logging.info(f"Sheet '{sheet_name}': found {len(sheet_connections)} connections with {len(df.columns)} columns")
                    
                except Exception as e:
                    logging.warning(f"Error reading sheet '{sheet_name}': {e}")
                    continue
            
            if sheets_processed == 0:
                logging.warning(f"No valid sheets found in QC file: {qc_file_path}")
                self._active = False
                return
            
            # Process all connections from all sheets
            self.connections.clear()
            self.ordered_connections.clear()
            self.original_ordered_connections.clear()
            self.qc_data_rows.clear()
            self.qc_scids.clear()
            
            for i, (from_pole_orig, to_pole_orig) in enumerate(all_connections):
                # Store original format (EXACT from QC file)
                original_connection = (from_pole_orig, to_pole_orig)
                self.original_ordered_connections.append(original_connection)
                
                # Store complete row data
                if i < len(all_data_rows):
                    self.qc_data_rows.append(all_data_rows[i])
                
                # Normalize SCIDs for matching purposes
                from_pole_norm = self._normalize_scid(from_pole_orig)
                to_pole_norm = self._normalize_scid(to_pole_orig)
                
                # Add normalized versions for internal matching
                normalized_connection = (from_pole_norm, to_pole_norm)
                self.connections.add(normalized_connection)
                self.connections.add((to_pole_norm, from_pole_norm))  # Add reverse for bidirectional matching
                self.ordered_connections.append(normalized_connection)
                
                self.qc_scids.add(from_pole_norm)
                self.qc_scids.add(to_pole_norm)
            
            self._active = len(self.connections) > 0
            self.qc_file_path = qc_file_path
            
            logging.info(f"Loaded QC file: {len(self.ordered_connections)} connections from {sheets_processed} sheets, {len(self.qc_scids)} unique SCIDs")
            
        except Exception as e:
            logging.error(f"Error loading QC file {qc_file_path}: {e}")
            self._active = False
    
    def _normalize_scid(self, scid):
        """
        Normalize SCID format with flexible extraction from entries with additional text
        Uses ignore_scid_keywords to remove specified keywords before normalizing
        
        Examples:
        - "023 AT&T" -> "23" (if "AT&T" in ignore keywords)
        - "178A Foreign Pole" -> "178A" (if "Foreign Pole" in ignore keywords)
        - "001A Unknown" -> "1A" (if "Unknown" in ignore keywords)
        - "023" -> "23"
        
        Args:
            scid (str): Raw SCID string
            
        Returns:
            str: Normalized SCID
        """
        scid = str(scid).strip()
        
        if not scid:
            return scid
        
        import re
        
        # Handle numeric SCIDs with leading zeros (simple case)
        if scid.isdigit():
            return str(int(scid))
        
        # Strategy 1: Remove ignore keywords from the SCID string
        scid_cleaned = scid
        for keyword in self.ignore_scid_keywords:
            if keyword.strip():  # Only process non-empty keywords
                # Create a case-insensitive pattern to match the keyword
                # Use word boundaries to avoid partial matches
                pattern = r'\b' + re.escape(keyword.strip()) + r'\b'
                scid_cleaned = re.sub(pattern, '', scid_cleaned, flags=re.IGNORECASE).strip()
        
        # Remove extra whitespace that might result from keyword removal
        scid_cleaned = re.sub(r'\s+', ' ', scid_cleaned).strip()
        
        # Strategy 2: Extract SCID pattern from the cleaned string
        # Look for patterns like:
        # - "023" -> extract "023"
        # - "178A" -> extract "178A" 
        # - "001A" -> extract "001A"
        
        # If after cleaning keywords, we have just a simple SCID, process it
        if scid_cleaned.isdigit():
            normalized = str(int(scid_cleaned))
            if scid != scid_cleaned:  # Log only if keywords were actually removed
                logging.info(f"QC SCID flexible extraction: '{scid}' -> cleaned '{scid_cleaned}' -> normalized '{normalized}'")
            return normalized
        
        # Pattern 1: Number + optional letter(s) at the beginning
        # Examples: "023A", "178A"
        match = re.match(r'^(\d+[A-Za-z]*)(?:\s+.*)?$', scid_cleaned)
        if match:
            base_scid = match.group(1)
            # Normalize the extracted base SCID (remove leading zeros)
            num_match = re.match(r'^(\d+)([A-Za-z]*)$', base_scid)
            if num_match:
                num_part = str(int(num_match.group(1)))
                alpha_part = num_match.group(2).upper()
                normalized = num_part + alpha_part
                if scid != scid_cleaned:  # Log only if keywords were actually removed
                    logging.info(f"QC SCID flexible extraction: '{scid}' -> cleaned '{scid_cleaned}' -> extracted '{base_scid}' -> normalized '{normalized}'")
                return normalized
        
        # Strategy 3: Fall back to original logic for edge cases
        # Pattern 2: Number + optional letter(s) + space + anything else
        # Examples: "023 remaining text", "178A remaining text"
        match = re.match(r'^(\d+[A-Za-z]*)\s+.*', scid)
        if match:
            base_scid = match.group(1)
            # Normalize the extracted base SCID (remove leading zeros)
            num_match = re.match(r'^(\d+)([A-Za-z]*)$', base_scid)
            if num_match:
                num_part = str(int(num_match.group(1)))
                alpha_part = num_match.group(2).upper()
                normalized = num_part + alpha_part
                logging.info(f"QC SCID flexible extraction: '{scid}' -> extracted '{base_scid}' -> normalized '{normalized}'")
                return normalized
        
        # Pattern 3: Simple alphanumeric SCID without spaces (fallback)
        # Examples: "001A", "023", "178A"
        match = re.match(r'^(\d+)([A-Za-z]*)$', scid_cleaned)
        if match:
            num_part = str(int(match.group(1)))
            alpha_part = match.group(2).upper()
            normalized = num_part + alpha_part
            if scid != scid_cleaned:  # Log only if keywords were actually removed
                logging.info(f"QC SCID flexible extraction: '{scid}' -> cleaned '{scid_cleaned}' -> normalized '{normalized}'")
            else:
                logging.debug(f"QC SCID normalization: '{scid}' -> '{normalized}'")
            return normalized
        
        # Pattern 4: Complex SCID with multiple parts - take the first numeric+alpha part
        # Examples: "118 MISM013", "023A Something Else"
        parts = scid_cleaned.split()
        if parts:
            first_part = parts[0]
            # Check if first part is a valid SCID pattern
            match = re.match(r'^(\d+)([A-Za-z]*)$', first_part)
            if match:
                num_part = str(int(match.group(1)))
                alpha_part = match.group(2).upper()
                normalized = num_part + alpha_part
                if scid != scid_cleaned:  # Log only if keywords were actually removed
                    logging.info(f"QC SCID flexible extraction: '{scid}' -> cleaned '{scid_cleaned}' -> first part '{first_part}' -> normalized '{normalized}'")
                else:
                    logging.debug(f"QC SCID extraction from parts: '{scid}' -> first part '{first_part}' -> normalized '{normalized}'")
                return normalized
        
        # If no patterns match, return as-is (but log it for debugging)
        if scid != scid_cleaned:  # Log only if keywords were actually removed
            logging.warning(f"QC SCID no pattern match after keyword removal: '{scid}' -> cleaned '{scid_cleaned}' -> returned as-is")
        else:
            logging.debug(f"QC SCID no pattern match: '{scid}' -> returned as-is")
        return scid_cleaned if scid != scid_cleaned else scid
    
    def is_active(self):
        """
        Check if QC reader is active (has loaded QC data)
        
        Returns:
            bool: True if QC data is loaded and active
        """
        return self._active
    
    def get_qc_scids(self):
        """
        Get set of all SCIDs mentioned in QC file
        
        Returns:
            set: Set of SCID strings
        """
        return self.qc_scids.copy()
    
    def get_ordered_connections(self):
        """
        Get connections in the order they appear in QC file (normalized for matching)
        
        Returns:
            list: List of (from_scid, to_scid) tuples in QC file order (normalized)
        """
        return self.ordered_connections.copy()
    
    def get_original_ordered_connections(self):
        """
        Get connections in EXACT format from QC file (preserving original format)
        
        Returns:
            list: List of (from_scid, to_scid) tuples in exact QC file format
        """
        return self.original_ordered_connections.copy()
    
    def get_qc_data_rows(self):
        """
        Get complete row data from QC file
        
        Returns:
            list: List of dictionaries containing complete row data from QC file
        """
        return self.qc_data_rows.copy()
    
    def has_connection(self, from_scid, to_scid):
        """
        Check if a specific connection exists in QC file
        
        Args:
            from_scid (str): Source SCID
            to_scid (str): Destination SCID
            
        Returns:
            bool: True if connection exists in QC file
        """
        if not self._active:
            return False
        
        # Normalize inputs
        from_scid = self._normalize_scid(str(from_scid))
        to_scid = self._normalize_scid(str(to_scid))
        
        return (from_scid, to_scid) in self.connections
    
    def get_connections_set(self):
        """
        Get set of all connections (bidirectional)
        
        Returns:
            set: Set of (from_scid, to_scid) tuples
        """
        return self.connections.copy()
    
    def get_all_connections(self):
        """
        Get all connections (alias for get_ordered_connections for compatibility)
        
        Returns:
            list: List of (from_scid, to_scid) tuples in QC file order (normalized)
        """
        return self.get_ordered_connections()
    
    def create_consolidated_qc_sheet(self, output_path=None):
        """
        Create a consolidated QC sheet that combines data from all sheets
        
        Args:
            output_path (str, optional): Path for the output file. If None, overwrites original file.
            
        Returns:
            str: Path to the created consolidated QC file
        """
        if not self._active or not self.qc_file_path:
            logging.warning("No QC file loaded - cannot create consolidated sheet")
            return None
            
        try:
            import openpyxl
            from pathlib import Path
            
            # Determine output path
            if output_path is None:
                original_path = Path(self.qc_file_path)
                output_path = original_path.parent / f"{original_path.stem}_Consolidated{original_path.suffix}"
            
            # Load the original workbook
            wb = openpyxl.load_workbook(self.qc_file_path, data_only=True)
            
            # Create new workbook for consolidated data
            new_wb = openpyxl.Workbook()
            
            # Remove default sheet and create QC sheet
            new_wb.remove(new_wb.active)
            qc_sheet = new_wb.create_sheet(title="QC")
            
            # Add header in row 3
            qc_sheet['A3'] = 'Pole'
            qc_sheet['B3'] = 'To Pole'
            
            # Style the header
            from openpyxl.styles import Font, Alignment
            header_font = Font(bold=True)
            qc_sheet['A3'].font = header_font
            qc_sheet['B3'].font = header_font
            qc_sheet['A3'].alignment = Alignment(horizontal='center')
            qc_sheet['B3'].alignment = Alignment(horizontal='center')
            
            # Use the connections that were already loaded and processed
            all_connections = self.get_original_ordered_connections()
            logging.info(f"Using {len(all_connections)} connections already loaded from QC file")
            
            # Write all connections to the QC sheet starting from row 4
            row_num = 4
            for from_pole, to_pole in all_connections:
                qc_sheet[f'A{row_num}'] = from_pole
                qc_sheet[f'B{row_num}'] = to_pole
                row_num += 1
            
            # Adjust column widths
            qc_sheet.column_dimensions['A'].width = 15
            qc_sheet.column_dimensions['B'].width = 15
            
            # Save the consolidated workbook
            new_wb.save(output_path)
            logging.info(f"Created consolidated QC sheet with {len(all_connections)} connections")
            logging.info(f"Consolidated QC file saved to: {output_path}")
            
            return str(output_path)
            
        except Exception as e:
            logging.error(f"Error creating consolidated QC sheet: {e}")
            return None
    
    def get_qc_span_length(self, from_scid_excel, to_scid_excel):
        """
        Get span length from QC data for a specific connection
        
        Args:
            from_scid_excel (str): From SCID as it appears in Excel (normalized)
            to_scid_excel (str): To SCID as it appears in Excel (normalized)
            
        Returns:
            str: Span length value from QC file, or empty string if not found
        """
        if not self._active:
            logging.debug(f"QC reader not active for span length lookup: {from_scid_excel} -> {to_scid_excel}")
            return ''
        
        logging.debug(f"Looking for QC span length: Excel SCIDs {from_scid_excel} -> {to_scid_excel}")
        
        # Normalize the Excel SCIDs for comparison
        from_normalized = self._normalize_scid(str(from_scid_excel))
        to_normalized = self._normalize_scid(str(to_scid_excel))
        
        logging.debug(f"Normalized Excel SCIDs: {from_normalized} -> {to_normalized}")
        
        # Look for the connection in normalized ordered connections
        for i, (qc_from_norm, qc_to_norm) in enumerate(self.ordered_connections):
            if qc_from_norm == from_normalized and qc_to_norm == to_normalized:
                # Found the connection, get the corresponding row data
                if i < len(self.qc_data_rows):
                    row_data = self.qc_data_rows[i]
                    qc_from_orig, qc_to_orig = self.original_ordered_connections[i]
                    
                    logging.debug(f"Found QC connection match: Excel {from_scid_excel}->{to_scid_excel} matches QC {qc_from_orig}->{qc_to_orig}")
                    logging.debug(f"Available QC columns: {list(row_data.keys())}")
                    
                    # Look for span length in common column names
                    span_columns = [
                        'Pole to Pole Span Length (from starting point)',
                        'Span Length',
                        'Pole to Pole Span Length',
                        'Distance',
                        'Span Distance'
                    ]
                    for col in span_columns:
                        if col in row_data:
                            value = str(row_data[col]).strip() if row_data[col] else ''
                            logging.debug(f"QC column '{col}' has value: '{value}'")
                            if value and value.lower() not in ['nan', 'none', '']:
                                logging.info(f"Found QC span length for {from_scid_excel} -> {to_scid_excel}: {value} (from QC {qc_from_orig} -> {qc_to_orig})")
                                return value
                    
                    logging.debug(f"No span length found in QC data for {from_scid_excel} -> {to_scid_excel}")
                break
        else:
            logging.debug(f"QC connection {from_scid_excel} -> {to_scid_excel} (normalized: {from_normalized} -> {to_normalized}) not found in QC connections")
        
        return ''
