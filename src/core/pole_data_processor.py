import pandas as pd
import logging
from pathlib import Path
import re
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

from .utils import Utils
from .config_manager import ConfigManager
from .qc_reader import QCReader
from .tension_calculator_com import TensionCalculatorCOM


class PoleDataProcessor:
    """Handles pole data processing and Excel output"""
    
    def __init__(self, config, geocoder=None, mapping_data=None, attachment_reader=None, qc_reader=None):
        self.config = config
        self.geocoder = geocoder
        self.mapping_data = mapping_data or []
        self.attachment_reader = attachment_reader
        self.qc_reader = qc_reader
        # Initialize tension calculator with configuration
        tension_config = self.config.get("tension_calculator", {})
        calculator_file_path = tension_config.get("file_path", "")
        worksheet_name = tension_config.get("worksheet_name", "Calculations")
        
        # Try COM-based calculator first, fall back to openpyxl-based if it fails
        try:
            self.tension_calculator = TensionCalculatorCOM(calculator_file_path, worksheet_name)
            # Test if it can initialize
            if calculator_file_path:
                test_result = self.tension_calculator.calculate_tension(100.0, 26.33, 25.0)
                if test_result is None:
                    logging.warning("COM-based tension calculator failed to initialize, trying openpyxl fallback")
                    from .tension_calculator import TensionCalculator
                    self.tension_calculator = TensionCalculator(calculator_file_path, worksheet_name)
                else:
                    logging.info("COM-based tension calculator initialized successfully")
            else:
                logging.info("No tension calculator file specified, tension calculation will be skipped")
        except Exception as e:
            logging.warning(f"COM-based tension calculator failed: {e}, trying openpyxl fallback")
            try:
                from .tension_calculator import TensionCalculator
                self.tension_calculator = TensionCalculator(calculator_file_path, worksheet_name)
                logging.info("Openpyxl-based tension calculator initialized successfully")
            except Exception as e2:
                logging.error(f"Both tension calculators failed: {e2}")
                # Create a dummy calculator that always returns None
                class DummyTensionCalculator:
                    def calculate_tension(self, *args, **kwargs):
                        return None
                    def cleanup(self):
                        pass
                self.tension_calculator = DummyTensionCalculator()
    
    def process_data(self, nodes_df, connections_df, sections_df, progress_callback=None, 
                    manual_routes=None, clear_existing_routes=False):
        """Process pole data"""
        if progress_callback:
            progress_callback(40, "Filtering pole data...")
        
        # Cache connections DataFrame for alternative section lookup
        self.connections_df_cache = connections_df
        
        # Normalize SCIDs and filter nodes
        nodes_df = nodes_df.copy()
        ignore_keywords = self.config.get('ignore_scid_keywords', [])
        nodes_df['scid'] = nodes_df['scid'].apply(lambda x: Utils.normalize_scid(x, ignore_keywords))
        nodes_df = nodes_df.drop_duplicates(subset='scid')
        
        # Sort nodes by SCID numerically
        nodes_df['sort_key'] = nodes_df['scid'].apply(Utils.extract_numeric_part)
        nodes_df = nodes_df.sort_values(by='sort_key')
        nodes_df = nodes_df.drop('sort_key', axis=1)
        
        # Filter valid SCIDs: node_type = 'pole' OR 'reference' AND pole_status != 'underground'
        filtered = Utils.filter_valid_nodes(nodes_df)
        
        if filtered.empty:
            raise ValueError("No valid pole or reference data found")
        
        # Log the filtered data for debugging
        poles_count = len(filtered[filtered['node_type'].str.strip().str.lower().eq('pole')])
        references_count = len(filtered[filtered['node_type'].str.strip().str.lower().eq('reference')])
        logging.info(f"Found {poles_count} valid poles and {references_count} valid references")
        
        # Create mappings
        mappings = self._create_mappings(nodes_df, filtered)
        
        if progress_callback:
            if not progress_callback(50, "Building connections..."):
                return []  # Stop processing if requested
        
        # Build temp rows
        temp_rows = self._build_temp_rows(connections_df, mappings, manual_routes, clear_existing_routes)
        
        # If QC file is active, only keep poles mentioned in QC file
        if self.qc_reader and self.qc_reader.is_active():
            qc_scids = self.qc_reader.get_qc_scids()
            original_count = len(temp_rows)
            temp_rows = {scid: data for scid, data in temp_rows.items() if scid in qc_scids}
            logging.info(f"QC filtering: reduced from {original_count} to {len(temp_rows)} poles (only QC SCIDs)")
            
            if not temp_rows:
                logging.warning("No poles found after QC filtering - check that QC SCIDs match pole SCIDs in data")
        
        if progress_callback:
            if not progress_callback(70, "Processing connections..."):
                return []  # Stop processing if requested

        # Process connections to generate output rows (one row per connection involving a pole)
        result_data = []
        
        # If QC file is active, generate output based on QC connections only
        if self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - filtering output to QC connections only")
            result_data = self._process_qc_filtered_connections(
                connections_df, mappings, sections_df
            )
        else:
            # Process each connection and generate rows (optimized logic)
            result_data = self._process_standard_connections(connections_df, mappings, sections_df)
        
        # Filter results based on manual routes if specified
        if manual_routes:
            manual_scids = {scid for route in manual_routes for scid in route['poles']}
            logging.info(f"Filtering results to manual route SCIDs: {sorted(manual_scids)}")
            
            original_count = len(result_data)
            result_data = [row for row in result_data if row.get('Pole') in manual_scids]
            logging.info(f"Manual route filtering: reduced from {original_count} to {len(result_data)} rows")
        
        # Sort result by pole SCID (unless QC file is active, then preserve QC order)
        if self.qc_reader and self.qc_reader.is_active():
            # QC file is active - result_data is already in QC order, don't sort
            logging.info("Preserving QC file order for output")
        else:
            # No QC file - sort by pole SCID as usual
            result_data.sort(key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))
        
        if progress_callback:
            if not progress_callback(90, f"Generated {len(result_data)} output rows"):
                return []  # Stop processing if requested
        
        return result_data
    

    
    def _create_mappings(self, nodes_df, filtered):
        """Create various lookup mappings"""
        return {
            'node_id_to_scid': nodes_df.set_index('node_id')['scid'].to_dict(),
            'scid_to_row': nodes_df.set_index('scid').to_dict('index'),
            'node_id_to_row': nodes_df.set_index('node_id').to_dict('index'),
            'valid_poles': set(filtered['node_id'])
        }
    
    def _process_standard_connections(self, connections_df, mappings, sections_df):
        """Process standard connections without QC filtering (optimized)"""
        result_data = []
        processed_connections = set()
        
        # Pre-filter connections to only valid poles for better performance
        valid_poles = mappings['valid_poles']
        mask = (connections_df['node_id_1'].isin(valid_poles)) & (connections_df['node_id_2'].isin(valid_poles))
        valid_connections = connections_df[mask]
        
        logging.info(f"Processing {len(valid_connections)} valid connections out of {len(connections_df)} total connections")
        
        for _, conn in valid_connections.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            
            connection_key = tuple(sorted([n1, n2]))
            if connection_key not in processed_connections:
                processed_connections.add(connection_key)
                
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                
                node1_data = mappings['node_id_to_row'].get(n1, {})
                node2_data = mappings['node_id_to_row'].get(n2, {})
                node1_type = str(node1_data.get('node_type', '')).strip().lower()
                node2_type = str(node2_data.get('node_type', '')).strip().lower()
                
                conn_info = {
                    'connection_id': conn.get('connection_id', ''),
                    'span_distance': conn.get('span_distance', '')
                }
                
                # Generate row(s) for this connection
                if node1_type == 'pole' and node2_type == 'reference':
                    # Pole -> Reference: Pole in "Pole" column, Reference in "To Pole" column
                    row_data = self._create_output_row(scid1, scid2, conn_info, node1_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        result_data.append(row_data)
                        logging.debug(f"Added pole->reference connection: {scid1} -> {scid2}")
                elif node1_type == 'reference' and node2_type == 'pole':
                    # Reference -> Pole: Pole in "Pole" column, Reference in "To Pole" column
                    row_data = self._create_output_row(scid2, scid1, conn_info, node2_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        result_data.append(row_data)
                        logging.debug(f"Added reference->pole connection: {scid2} -> {scid1}")
                elif node1_type == 'pole' and node2_type == 'pole':
                    # Pole -> Pole: First pole in "Pole" column, Second pole in "To Pole" column
                    row_data = self._create_output_row(scid1, scid2, conn_info, node1_data, mappings['scid_to_row'], sections_df)
                    if row_data:
                        result_data.append(row_data)
                        logging.debug(f"Added pole->pole connection: {scid1} -> {scid2}")
        
        # Count pole-to-reference connections for logging
        pole_ref_count = sum(1 for row in result_data if row.get('To Pole', '') and 
                           any(ref_scid in row.get('To Pole', '') for ref_scid in 
                               [scid for scid, data in mappings['scid_to_row'].items() 
                                if str(data.get('node_type', '')).strip().lower() == 'reference']))
        
        logging.info(f"Generated {len(result_data)} total connections, including {pole_ref_count} pole-to-reference connections")
        
        return result_data
    
    def _build_temp_rows(self, connections_df, mappings, manual_routes, clear_existing_routes):
        """Build temporary rows for processing"""
        temp = {}
        processed = set()
        
        # Initialize all valid poles
        for node_id in mappings['valid_poles']:
            scid = mappings['node_id_to_scid'][node_id]
            node_data = mappings['node_id_to_row'].get(node_id, {})
            guy_info = self._extract_guy_info(node_data.get('mr_note', ''))
            
            temp[scid] = {
                'Pole': scid,
                'Guy Size': '',
                'Guy Lead': ', '.join(guy_info['leads']),
                'Guy Direction': ', '.join(guy_info['directions']),
                'To Pole': '',
                'connection_id': '',
                'span_distance': ''
            }
        
        # Skip Excel connection processing if QC file is active
        if self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - skipping Excel connection processing")
            connection_data = {}
        else:
            # Process Excel connections
            connection_data = self._process_excel_connections(
                connections_df, mappings, temp, processed, clear_existing_routes
            )
        
        # Apply manual routes (only if QC file is not active)
        if manual_routes and not (self.qc_reader and self.qc_reader.is_active()):
            self._apply_manual_routes(manual_routes, temp, connection_data)
        elif manual_routes and self.qc_reader and self.qc_reader.is_active():
            logging.info("QC file is active - manual routes will be ignored in favor of QC connections")
        
        logging.info(f"Built {len(temp)} pole records with routing information")
        return temp

    def _process_excel_connections(self, connections_df, mappings, temp, processed, clear_existing_routes):
        """Process connections from Excel data with enhanced reference node logic"""
        logging.info("Processing automatic connections from Excel data...")
        connection_data = {}
        
        for _, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            
            if (n1 in mappings['valid_poles'] and n2 in mappings['valid_poles']):
                connection_key = tuple(sorted([n1, n2]))
                if connection_key not in processed:
                    processed.add(connection_key)
                    scid1 = mappings['node_id_to_scid'][n1]
                    scid2 = mappings['node_id_to_scid'][n2]
                    
                    # Get node types to handle reference nodes correctly
                    node1_data = mappings['node_id_to_row'].get(n1, {})
                    node2_data = mappings['node_id_to_row'].get(n2, {})
                    node1_type = str(node1_data.get('node_type', '')).strip().lower()
                    node2_type = str(node2_data.get('node_type', '')).strip().lower()
                    
                    conn_info = {
                        'connection_id': conn.get('connection_id', ''),
                        'span_distance': conn.get('span_distance', '')
                    }
                    
                    # Store connection data (use sorted tuple as key to avoid duplication)
                    connection_key = tuple(sorted([scid1, scid2]))
                    connection_data[connection_key] = conn_info
                    
                    if not clear_existing_routes:
                        # Handle reference node logic: references must be at 'To Pole'
                        if node2_type == 'reference' and node1_type == 'pole':
                            # scid1 is pole, scid2 is reference
                            temp[scid1].update({'To Pole': scid2, **conn_info})
                        elif node1_type == 'reference' and node2_type == 'pole':
                            # scid2 is pole, scid1 is reference  
                            temp[scid2].update({'To Pole': scid1, **conn_info})
                        elif node1_type == 'pole' and node2_type == 'pole':
                            # Both are poles, use normal connection logic
                            temp[scid1].update({'To Pole': scid2, **conn_info})
                        else:
                            # Default behavior for other cases
                            temp[scid1].update({'To Pole': scid2, **conn_info})
        
        if clear_existing_routes:
            logging.info("Cleared existing route data as requested")
            for scid in temp:
                temp[scid]['To Pole'] = ''
        
        return connection_data

    def _apply_manual_routes(self, manual_routes, temp, connection_data):
        """Apply manual routes to pole data"""
        logging.info(f"Applying {len(manual_routes)} manual routes...")
        
        # Verify that every defined connection actually exists in the Excel data
        missing = []
        for route in manual_routes:
            for from_scid, to_scid in route['connections']:
                connection_key = tuple(sorted([from_scid, to_scid]))
                if from_scid not in temp or to_scid not in temp or connection_key not in connection_data:
                    missing.append((from_scid, to_scid))
        
        if missing:
            msg = "The following manual-route connections are invalid or missing:\n" + \
                  "\n".join(f"{a} → {b}" for a, b in missing)
            raise ValueError("Invalid manual routes detected")
        
        for route_idx, route in enumerate(manual_routes):
            logging.info(f"Processing manual route {route_idx + 1}: {' → '.join(route['poles'])}")
            for from_scid, to_scid in route['connections']:
                connection_key = tuple(sorted([from_scid, to_scid]))
                conn_info = connection_data.get(connection_key, {})
                
                temp[from_scid].update({
                    'To Pole': to_scid,
                    'connection_id': conn_info.get('connection_id', ''),
                    'span_distance': conn_info.get('span_distance', '')
                })
                logging.info(f"  Set {from_scid} → {to_scid}")
                
                if not conn_info.get('connection_id'):
                    logging.warning(f"  No Excel connection data found for {from_scid} → {to_scid}")
            
            # Handle dead-end
            last_pole = route['poles'][-1]
            if last_pole in temp:
                temp[last_pole]['To Pole'] = ''
                logging.info(f"  Set {last_pole} as dead-end")
    
    def _extract_guy_info(self, note):
        """Extract guy information from notes.
           Supports multiple formats:
             - "PL NEW SINGLE HELIX ANCHOR 15' S WITH OFFSET" -> Guy Lead = "15'", Guy Direction = "S"
             - "PL NEW xxxxxx ANCHOR 20'6" NW" -> Guy Lead = "20'6"", Guy Direction = "NW"
             - "ANCHOR 10' W"  -> Guy Lead = "10'" and Guy Direction = "W"
             - "ANCHOR 15'6" NW" -> Guy Lead = "15'6"" and Guy Direction = "NW"
             - "GUY 3/8" EHS 20' S" -> Guy Size = "3/8" EHS", Guy Lead = "20'", Guy Direction = "S"
             - "5/16" EHS GUY 15' N" -> Guy Size = "5/16" EHS", Guy Lead = "15'", Guy Direction = "N"
        """
        if not note or pd.isna(note):
            return {'leads': [], 'directions': [], 'sizes': []}
        
        note = str(note).upper()
        leads = []
        directions = []
        sizes = []
        
        # Pattern 0: PL NEW format - "PL NEW SINGLE HELIX ANCHOR 15' S WITH OFFSET"
        pl_new_pattern = r"PL\s+NEW\s+[A-Z\s]+\s+ANCHOR\s+(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})(?:\s|$)"
        pl_new_matches = re.findall(pl_new_pattern, note)
        
        # If PL NEW patterns are found, only use those and skip other patterns
        if pl_new_matches:
            for feet, inches, direction in pl_new_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in PL NEW format
        else:
            # Only process other patterns if no PL NEW patterns were found
            
            # Pattern 1: ANCHOR format - "ANCHOR 10' W"
            anchor_pattern = r"ANCHOR\s+(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})"
            anchor_matches = re.findall(anchor_pattern, note)
            for feet, inches, direction in anchor_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in ANCHOR format
            
            # Pattern 2: GUY with size - "GUY 3/8" EHS 20' S" or "5/16" EHS GUY 15' N"
            guy_pattern = r"(?:GUY\s+)?(\d+/\d+\"\s*EHS|[\d.]+\"\s*EHS)\s*(?:GUY\s+)?(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})"
            guy_matches = re.findall(guy_pattern, note)
            for size, feet, inches, direction in guy_matches:
                # Build Guy Lead string preserving inches if provided
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                size = size.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append(size)
            
            # Pattern 3: General guy pattern - any remaining patterns with just lead/direction
            # Make this more restrictive to avoid matching height values
            general_pattern = r"(?:^|\s)(\d+)'(?:\s*(\d+)\")?\s+([NSEW]{1,2})(?:\s|$)"
            general_matches = re.findall(general_pattern, note)
            for feet, inches, direction in general_matches:
                if inches:
                    lead = f"{feet}'{inches}\""
                else:
                    lead = f"{feet}'"
                lead = lead.strip()
                direction = direction.strip()
                combined = f"{lead} {direction}"
                if combined not in [f"{l} {d}" for l, d in zip(leads, directions)]:
                    leads.append(lead)
                    directions.append(direction)
                    sizes.append('')  # No size info in general format
        
        return {'leads': leads, 'directions': directions, 'sizes': sizes}

    def _create_output_row(self, pole_scid, to_pole_scid, conn_info, pole_node_data, scid_to_row, sections_df):
        """Create an output row for a connection involving a pole"""
        try:
            # Get pole data from the pole_scid (this should always be a pole, not a reference)
            node = scid_to_row.get(pole_scid, pole_node_data)
            
            # Determine connection type by checking if to_pole_scid is a reference
            to_pole_data = scid_to_row.get(to_pole_scid, {})
            to_pole_type = str(to_pole_data.get('node_type', '')).strip().lower()
            is_pole_to_reference = (to_pole_type == 'reference')
            
            logging.debug(f"Connection {pole_scid} -> {to_pole_scid}: to_pole_type='{to_pole_type}', is_pole_to_reference={is_pole_to_reference}")
            
            # Find section data for this connection
            connection_id = conn_info.get('connection_id', '')
            section = self._find_section(connection_id, sections_df, pole_scid, to_pole_scid)
            logging.debug(f"Section lookup for {pole_scid} -> {to_pole_scid}: connection_id='{connection_id}', section found: {section is not None}")
            
            # If no section found via connection_id, try alternative approach for pole-to-reference connections
            if section is None and sections_df is not None and not sections_df.empty:
                # Get the pole's node_id for connection lookup
                pole_node_id = node.get('node_id', '') if node else ''
                
                if pole_node_id and hasattr(self, 'connections_df_cache'):
                    # Look for connections that involve this pole's node_id
                    potential_connection_ids = []
                    
                    for _, conn_row in self.connections_df_cache.iterrows():
                        conn_node1 = str(conn_row.get('node_id_1', '')).strip()
                        conn_node2 = str(conn_row.get('node_id_2', '')).strip()
                        
                        if pole_node_id == conn_node1 or pole_node_id == conn_node2:
                            potential_connection_ids.append(conn_row.get('connection_id', ''))
                    
                    # Try to find sections using these connection_ids
                    if potential_connection_ids:
                        for conn_id in potential_connection_ids:
                            potential_section = self._find_section(conn_id, sections_df, pole_scid, to_pole_scid)
                            if potential_section is not None:
                                section = potential_section
                                logging.debug(f"Using alternative section lookup for {pole_scid} -> {to_pole_scid}: found section data via connection {conn_id}")
                                break
                    
                    if section is None:
                        logging.debug(f"No sections found for pole {pole_scid} (node_id: {pole_node_id}) via alternative lookup")
                else:
                    logging.debug(f"Alternative section lookup skipped for {pole_scid} -> {to_pole_scid}: no connections_df_cache or node_id")
            
            # Create empty section if none found
            if section is None:
                section = pd.Series()
            
            # Get mapped elements for processing
            mapped_elements = self._get_mapped_elements()
            
            # Process attachments and midspan data - pass connection type info
            result = self._process_attachments(node, section, mapped_elements, pole_scid, is_pole_to_reference)
            
            # Add basic connection information
            result['Pole'] = pole_scid
            result['To Pole'] = to_pole_scid
            
            # Get the initial span length and format it
            initial_span_distance = conn_info.get('span_distance', '')
            final_span_str = initial_span_distance

            # If QC is active, apply tolerance to get the final span length
            if self.qc_reader and self.qc_reader.is_active():
                tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                qc_span_str = self.qc_reader.get_qc_span_length(pole_scid, to_pole_scid)
                final_span_str = self._apply_span_length_tolerance(initial_span_distance, qc_span_str, tolerance)
                logging.info(f"Span tolerance applied for {pole_scid} -> {to_pole_scid}: "
                             f"Excel='{initial_span_distance}', QC='{qc_span_str}', Final='{final_span_str}'")

            result['Span Length'] = self._format_span_distance(final_span_str)
            
            # Use the final, adjusted span length for tension calculation
            span_length_for_tension = self._parse_span_length(final_span_str)

            # Add pole information
            result['Address'] = self._get_pole_address(node)
            result['Pole Height/Class'] = self._format_pole_height_class(node)
            result['Existing Risers'] = self._count_existing_risers(node)
            
            # Add coordinates if available
            if node and 'latitude' in node:
                result['Latitude'] = node['latitude']
            if node and 'longitude' in node:
                result['Longitude'] = node['longitude']
            
            # Add guy information - use same priority order as Notes field
            guy_note = ''
            if node:
                # Use mr_note only
                mr_note = node.get('mr_note', '')
                guy_note = str(mr_note) if pd.notna(mr_note) else ''

            guy_info = self._extract_guy_info(guy_note)
            result['Guy Info'] = guy_info
            
            # Add Map field (can be customized based on requirements)
            result['Map'] = ''  # Empty for now, can be populated based on business logic
            
            # Add Line No. (will be set during output writing)
            result['Line No.'] = ''
            
            # Extract individual telecom provider heights from All_Comm_Heights
            all_comm_heights = result.get('All_Comm_Heights', '')
            
            # Initialize telecom provider fields - preserve existing values from attachment processing
            if 'Proposed MetroNet' not in result:
                result['Proposed MetroNet'] = ''
            if 'Verizon' not in result:
                result['Verizon'] = ''
            if 'AT&T' not in result:
                result['AT&T'] = ''
            if 'Comcast' not in result:
                result['Comcast'] = ''
            if 'Zayo' not in result:
                result['Zayo'] = ''
            if 'Jackson ISD' not in result:
                result['Jackson ISD'] = ''
            
            # Parse individual heights from All_Comm_Heights - only if not already populated
            # Skip provider-specific parsing for pole-to-reference connections
            if all_comm_heights and not is_pole_to_reference:
                import re
                # Pattern to match "height (provider)" format
                height_pattern = r"(\d+'\s*\d*\"?)\s*\(([^)]+)\)"
                matches = re.findall(height_pattern, all_comm_heights)
                
                for height, provider in matches:
                    # Clean up height formatting - ensure proper format
                    height = height.strip()
                    
                    provider = provider.strip()
                    if 'MetroNet' in provider and not result['Proposed MetroNet']:
                        result['Proposed MetroNet'] = height
                    elif 'Verizon' in provider and not result['Verizon']:
                        result['Verizon'] = height
                    elif 'AT&T' in provider and not result['AT&T']:
                        result['AT&T'] = height
                    elif 'Comcast' in provider and not result['Comcast']:
                        result['Comcast'] = height
                    elif 'Zayo' in provider and not result['Zayo']:
                        result['Zayo'] = height
                    elif 'Jackson' in provider and not result['Jackson ISD']:
                        result['Jackson ISD'] = height
            elif all_comm_heights and is_pole_to_reference:
                # For pole-to-reference connections, skip provider-specific field population
                logging.debug(f"Skipping provider-specific field parsing for pole-to-reference connection (SCID {pole_scid})")
            
            # Calculate tensions for mapped providers
            tension_mappings = self._get_tension_mappings()

            # Check if calculator path is set and exists before attempting calculations
            calculator_path_str = self.config.get("tension_calculator", {}).get("file_path", "")
            can_calculate_tension = calculator_path_str and Path(calculator_path_str).exists()

            if tension_mappings:
                if can_calculate_tension:
                    for provider in tension_mappings:
                        logging.info(f"Checking tension calculation for provider: {provider}")
                        
                        if provider in result:  # Provider has attachment height
                            logging.info(f"Provider {provider} found in result data with value: {result[provider]}")
                            
                            # Use the final span length determined above
                            if span_length_for_tension:
                                # Get attachment height
                                attachment_height = self._parse_height_value(result.get(provider, ''))
                                logging.info(f"Attachment height for {provider}: {attachment_height} (raw: {result.get(provider, 'N/A')})")
                                
                                # Get midspan height from section sheet
                                midspan_height = None
                                section = self._find_section(conn_info.get('connection_id', ''), sections_df, pole_scid, to_pole_scid)
                                if section is not None:
                                    # Look for POA_ columns that contain MetroNet in the corresponding owner column
                                    for col in section.index:
                                        if col.startswith("POA_") and col.endswith("HT"):
                                            owner_col = col[:-2]  # Remove 'HT' suffix
                                            if owner_col in section.index:
                                                owner_value = str(section[owner_col]).lower()
                                                if any(keyword.lower() in owner_value for keyword in ["metronet", "metro"]):
                                                    midspan_value = section[col]
                                                    if pd.notna(midspan_value):
                                                        midspan_height = self._parse_height_value(midspan_value)
                                                        logging.info(f"Found midspan height from section sheet column {col}: {midspan_height} (raw: {midspan_value})")
                                                        break
                                
                                if midspan_height is None:
                                    logging.warning(f"No midspan height found in section sheet for {provider}")
                                
                                logging.info(f"Final heights for {provider}: attachment={attachment_height}, midspan={midspan_height}")
                                
                                if attachment_height is not None and midspan_height is not None:
                                    # Round height values to 2 decimal places
                                    attachment_height = round(float(attachment_height), 2)
                                    midspan_height = round(float(midspan_height), 2)
                                    
                                    logging.info(f"All data available for {provider} tension calculation: span={span_length_for_tension}, attach={attachment_height:.2f}, midspan={midspan_height:.2f}")
                                    
                                    # Build 2-decimal numeric strings for calculator inputs
                                    attachment_formatted = f"{attachment_height:.2f}"
                                    midspan_formatted = f"{midspan_height:.2f}"
        
                                    logging.info(
                                        f"Formatted values for tension calculation: span={span_length_for_tension:.2f}, "
                                        f"attach={attachment_formatted}, midspan={midspan_formatted}"
                                    )
        
                                    # Calculate tension using decimal-feet values
                                    tension = self.tension_calculator.calculate_tension(
                                        span_length_for_tension, attachment_formatted, midspan_formatted
                                    )
        
                                    if tension is not None:
                                        # Store whole-number tension in the mapped field
                                        tension_field = tension_mappings[provider]
                                        result[tension_field] = str(int(round(tension)))
                                        logging.info(
                                            f"✅ Calculated tension for {provider}: {int(round(tension))} lbs -> {tension_field}"
                                        )
                                    else:
                                        logging.warning(f"❌ Failed to calculate tension for {provider} - calculator returned None")
                                else:
                                    if attachment_height is None:
                                        logging.warning(f"❌ Missing attachment height for {provider}")
                                    if midspan_height is None:
                                        logging.warning(f"❌ Missing midspan height for {provider}")
                            else:
                                logging.warning(f"❌ No span length available for {provider} tension calculation")
                        else:
                            logging.warning(f"❌ Provider {provider} not found in result data")
                else:
                    logging.info("Tension calculation skipped: calculator file not specified or not found.")
            result['Cable Type 1'] = ''
            result['Cable Diameter 1'] = ''
            result['Cable Type 2'] = ''
            result['Cable Diameter 2'] = ''
            result['Total Bundle Diameter'] = ''
            
            # Add guy fields from guy_info
            if guy_info and (guy_info['leads'] or guy_info['directions']):
                result['Guy Size'] = ', '.join(guy_info['sizes']) if guy_info['sizes'] else ''
                result['Guy Lead'] = ', '.join(guy_info['leads']) if guy_info['leads'] else ''
                result['Guy Direction'] = ', '.join(guy_info['directions']) if guy_info['directions'] else ''
            else:
                result['Guy Size'] = ''
                result['Guy Lead'] = ''
                result['Guy Direction'] = ''
            
            # Add notes field - check multiple possible note fields
            notes = ''
            if node:
                # Use mr_note only
                mr_note = node.get('mr_note', '')
                notes = str(mr_note) if pd.notna(mr_note) else ''

            result['Notes'] = notes
            
            return result
            
        except Exception as e:
            logging.error(f"Error creating output row for {pole_scid} -> {to_pole_scid}: {e}")
            return None

    def _find_section(self, connection_id, sections_df, pole_scid=None, to_pole_scid=None):
        """Find section data for a connection_id, choosing section with lowest Proposed MetroNet height if multiple entries exist.
        If multiple rows match, further filter by pole_scid and to_pole_scid if columns exist."""
        if sections_df is None or sections_df.empty:
            return None
        
        # Filter for matching connection_id
        matching = sections_df[sections_df['connection_id'] == connection_id]
        
        # If possible, further filter by pole_scid and to_pole_scid
        pole_cols = [col for col in sections_df.columns if col.lower() in ['pole', 'from_pole', 'pole_scid', 'from_scid']]
        to_pole_cols = [col for col in sections_df.columns if col.lower() in ['to_pole', 'to_scid']]
        if pole_scid and to_pole_scid and not matching.empty:
            for pole_col in pole_cols:
                matching = matching[matching[pole_col] == pole_scid]
            for to_pole_col in to_pole_cols:
                matching = matching[matching[to_pole_col] == to_pole_scid]
        
        if matching.empty:
            return None
        
        if len(matching) == 1:
            return matching.iloc[0]
        
        # Choose entry with lowest overall attachment height when multiple entries exist
        if len(matching) > 1:
            matching_copy = matching.copy()
            
            # Find all height columns (POA_*HT)
            height_cols = [col for col in matching.columns if col.startswith("POA_") and col.endswith("HT")]
            
            if height_cols:
                # Calculate the lowest height for each row across all height columns
                min_heights = []
                for idx in matching.index:
                    row_heights = []
                    for ht_col in height_cols:
                        height_value = matching.loc[idx, ht_col]
                        if pd.notna(height_value):
                            try:
                                height_decimal = Utils.parse_height_decimal(height_value)
                                if height_decimal is not None:
                                    row_heights.append(height_decimal)
                            except:
                                continue
                    
                    # Find the minimum height for this row
                    min_height = min(row_heights) if row_heights else float('inf')
                    min_heights.append(min_height)
                
                # Find the row with the overall lowest height
                if min_heights and any(h != float('inf') for h in min_heights):
                    min_idx = min_heights.index(min(min_heights))
                    return matching.iloc[min_idx]
            
            # If no valid heights found, return first entry
            return matching.iloc[0]
        
        return matching.iloc[0]
    
    def _get_mapped_elements(self):
        """Get mapped providers and comm options from mapping data"""
        mapped = {
            'providers': set(),
            'midspan': set(),
            'comm_attach': set(),
            'comm_midspan': set()
        }
        
        for element, attribute, output in self.mapping_data:
            # Check if element is a telecom provider
            if element in self.config["telecom_providers"]:
                if attribute == "Attachment Ht":
                    mapped['providers'].add(element)
                elif attribute == "Midspan Ht":
                    mapped['midspan'].add(element)
            
            # Check if output is a comm field (separate check)
            if output in ["comm1", "comm2", "comm3", "comm4"]:
                if attribute == "Attachment Ht":
                    mapped['comm_attach'].add(output)
                elif attribute == "Midspan Ht":
                    mapped['comm_midspan'].add(output)
        
        return mapped
    
    def _process_attachments(self, node, section, mapped_elements, scid, is_pole_to_reference=False):
        """Process all attachment data for a pole"""
        # Initialize attachment dictionaries
        attach = {p: "" for p in mapped_elements['providers']}
        attach_midspan = {f"{p}_Midspan": "" for p in mapped_elements['midspan']}
        comm_attach = {c: "" for c in mapped_elements['comm_attach']}
        comm_midspan = {f"{c}_Midspan": "" for c in mapped_elements['comm_midspan']}
        
        # Add new fields for comprehensive communication data
        comm_attach['All_Comm_Heights'] = ""
        comm_attach['Total_Comm_Count'] = ""
        
        # Process attachment data from new file format
        telecom_heights = []
        power_heights = []
        power_midspan_heights = []
        all_telecom_attachments = []
        all_telecom_midspan = []
        
        # Get power and telecom attachments from attachment reader
        if self.attachment_reader:
            logging.debug(f"Processing attachments for pole {scid}")
            
            # Get power attachment (unchanged)
            power_attachment = self.attachment_reader.find_power_attachment(scid, self.config["power_keywords"])
            if power_attachment:
                power_heights.append((power_attachment['height_decimal'], power_attachment['height']))
                logging.info(f"Added power attachment for SCID {scid}: {power_attachment['height']}")
            else:
                logging.debug(f"No power attachment found for SCID {scid}")
            
            # Process ALL communication attachments from raw SCID data directly (unified approach)
            raw_scid_data = self.attachment_reader.get_scid_data(scid)
            if not raw_scid_data.empty:
                # Look for ALL communication-related entries using configurable keywords
                comm_keywords = self.config.get("comm_keywords", ['catv com', 'telco com', 'fiber optic com', 'insulator', 'power guy', 'communication', 'comm'])
                processed_attachments = {}  # Track by provider for provider-specific fields
                
                for _, row in raw_scid_data.iterrows():
                    measured = str(row.get('measured', '')).lower()
                    company = str(row.get('company', '')).lower()
                    
                    # Check if this is a communication attachment
                    is_comm = any(kw in measured for kw in comm_keywords) or any(kw in company for kw in comm_keywords)
                    
                    if is_comm and 'height_in_inches' in row:
                        try:
                            height_value = row['height_in_inches']
                            if pd.isna(height_value) or str(height_value).strip() == '':
                                continue
                                
                            # Clean and convert height value
                            height_str = str(height_value).replace('"', '').replace('″', '').strip()
                            try:
                                height_inches = float(pd.to_numeric(height_str, errors='coerce'))
                                if not pd.isna(height_inches) and height_inches > 0:
                                    height_formatted = Utils.inches_to_feet_format(str(int(height_inches)))
                                    if height_formatted:  # Only proceed if conversion was successful
                                        height_decimal = height_inches / 12
                                        
                                        # Determine provider for this attachment
                                        provider = None
                                        company_str = str(row.get('company', '')).strip()
                                        
                                        # Match to configured telecom providers
                                        for provider_name, keywords in self.config.get("telecom_keywords", {}).items():
                                            for keyword in keywords:
                                                if keyword.lower() in company.lower():
                                                    provider = provider_name
                                                    break
                                            if provider:
                                                break
                                        
                                        if not provider:
                                            provider = company_str if company_str else "Unknown"
                                        
                                        # Add to provider-specific processing (for provider fields)
                                        if not is_pole_to_reference and provider in mapped_elements['providers']:
                                            if provider not in processed_attachments:
                                                processed_attachments[provider] = []
                                            processed_attachments[provider].append((height_decimal, height_formatted))
                                        
                                        # Add to comprehensive list for All_Comm_Heights
                                        provider_info = f"{company_str} - {row.get('measured', 'Unknown')}" if company_str else row.get('measured', 'Unknown')
                                        entry = (height_decimal, height_formatted, provider_info)
                                        
                                        # Check if this exact height is already captured
                                        existing_heights = [x[0] for x in all_telecom_attachments]
                                        height_exists = any(abs(height_decimal - existing_height) < 0.01 for existing_height in existing_heights)
                                        
                                        if not height_exists:
                                            all_telecom_attachments.append(entry)
                                            telecom_heights.append(height_decimal)
                                            logging.debug(f"Added comm attachment for SCID {scid}: {height_formatted} ({provider_info})")
                                    else:
                                        logging.warning(f"Failed to format height {height_inches} for SCID {scid}")
                                else:
                                    logging.debug(f"Invalid height value {height_value} for SCID {scid}")
                            except Exception as e:
                                logging.warning(f"Error processing comm attachment for SCID {scid}: {e}")
                        except Exception as e:
                            logging.warning(f"Error processing comm attachment for SCID {scid}: {e}")
                
                # Process provider-specific attachments (combine multiple heights per provider)
                if not is_pole_to_reference:
                    for provider, height_list in processed_attachments.items():
                        if height_list:
                            # Sort heights from highest to lowest
                            height_list.sort(key=lambda x: x[0], reverse=True)
                            # Combine multiple heights for this provider
                            combined_heights = ', '.join([h[1] for h in height_list])
                            attach[provider] = combined_heights
                            logging.info(f"Set {provider} attachment for SCID {scid}: {combined_heights}")
                else:
                    logging.debug(f"Skipping provider-specific attachment processing for pole-to-reference connection (SCID {scid})")
        else:
            logging.warning("No attachment reader available - attachment data will not be processed")
        
        # Process section data for midspan - process for ALL connections (pole-to-pole AND pole-to-reference)
        midspan_processed_count = 0
        for col in section.index:
            if col.startswith("POA_") and not col.endswith("HT"):
                owner = str(section[col])
                ht_col = f"{col}HT"
                if ht_col in section and pd.notna(section[ht_col]):
                    fmt = Utils.parse_height_format(section[ht_col])
                    dec = Utils.parse_height_decimal(section[ht_col])
                    
                    self._process_midspan(owner, fmt, dec, attach_midspan, 
                                        power_midspan_heights, all_telecom_midspan, mapped_elements)
                    midspan_processed_count += 1
                    logging.debug(f"Processed midspan data for SCID {scid}, {col}: owner='{owner}', height={fmt}")
        
        if midspan_processed_count > 0:
            logging.debug(f"Total midspan entries processed for SCID {scid}: {midspan_processed_count}")
        else:
            logging.debug(f"No midspan data found in section for SCID {scid}")
        
        # Log connection type for debugging
        connection_type = "pole-to-reference" if is_pole_to_reference else "pole-to-pole"
        logging.debug(f"Processed {connection_type} connection (SCID {scid}): {midspan_processed_count} midspan entries")
        
        # Sort and assign comm attachments (now includes ALL communication heights)
        self._assign_comm_attachments(all_telecom_attachments, comm_attach, mapped_elements['comm_attach'])
        self._assign_comm_attachments(all_telecom_midspan, comm_midspan, mapped_elements['comm_midspan'])
        
        # Calculate power heights
        power_data = self._calculate_power_heights(power_heights, power_midspan_heights, telecom_heights)
        
        # Log final results for debugging
        if any(attach.values()) or any(power_data.values()) or comm_attach.get('All_Comm_Heights'):
            logging.info(f"Final attachments for SCID {scid}:")
            for key, value in {**attach, **power_data, **comm_attach}.items():
                if value and key in ['All_Comm_Heights', 'Total_Comm_Count']:
                    logging.info(f"  {key}: {value}")
        
        # Add streetlight (bottom of bracket) height
        streetlight_from_find = self.attachment_reader.find_streetlight_attachment(scid) if self.attachment_reader else None
        
        # New: Find street light for power company, measured contains 'street'
        street_light_height_processed = ""
        if self.attachment_reader:
            try:
                df_scid_data = self.attachment_reader.get_scid_data(scid)
                if not df_scid_data.empty:
                    power_company_config = self.config.get("power_company", "").strip().lower()

                    if power_company_config:
                        df_filtered = df_scid_data.copy()
                        df_filtered['company_stripped'] = df_filtered['company'].astype(str).str.strip().str.lower()
                        
                        power_company_pattern = r'\b' + re.escape(power_company_config) + r'\b'
                        power_company_rows = df_filtered[df_filtered['company_stripped'].str.contains(power_company_pattern, na=False, regex=True)]

                        if not power_company_rows.empty:
                            pc_rows_copy = power_company_rows.copy()
                            pc_rows_copy['measured_stripped'] = pc_rows_copy['measured'].astype(str).str.strip().str.lower()
                            
                            street_rows = pc_rows_copy[pc_rows_copy['measured_stripped'].str.contains('street', na=False)]

                            if not street_rows.empty:
                                s_rows_copy = street_rows.copy()
                                s_rows_copy['height_numeric'] = pd.to_numeric(
                                    s_rows_copy['height_in_inches'].astype(str).str.replace('"', '').str.replace('″', ''),
                                    errors='coerce'
                                )
                                s_rows_copy = s_rows_copy.dropna(subset=['height_numeric'])
                                if not s_rows_copy.empty:
                                    min_row = s_rows_copy.loc[s_rows_copy['height_numeric'].idxmin()]
                                    street_light_height_processed = Utils.inches_to_feet_format(str(int(min_row['height_numeric'])))
                    else:
                        logging.debug(f"SCID {scid}: Power company not configured. Skipping 'Street Light Height' processing.")
            except Exception as e:
                logging.error(f"Error processing street light height for SCID {scid}: {e}")
        
        result = {**attach, **attach_midspan, **comm_attach, **comm_midspan, **power_data}
        
        # This is for the pre-existing field 'Streetlight (bottom of bracket)'
        if streetlight_from_find:
            result['Streetlight (bottom of bracket)'] = streetlight_from_find['height']
        
        # This is for the new field 'Street Light Height'
        if street_light_height_processed:
            result['Street Light Height'] = street_light_height_processed
        
        return result
    
    def _process_midspan(self, owner, fmt, dec, attach_midspan, 
                        power_midspan_heights, all_telecom_midspan, mapped_elements):
        """Process midspan data"""
        # MetroNet midspan
        if self._match_metronet(owner):
            if "Proposed MetroNet" in mapped_elements['midspan']:
                attach_midspan["Proposed MetroNet_Midspan"] = fmt
        
        # Telecom midspan
        matched = self._match_telecom_provider(owner)
        if matched and matched in mapped_elements['midspan']:
            attach_midspan[f"{matched}_Midspan"] = fmt
        
        # Collect for comm sorting
        if matched or self._match_metronet(owner):
            if dec is not None and fmt:
                all_telecom_midspan.append((dec, fmt, owner))
        
        # Power midspan
        if any(kw.lower() in owner.lower() for kw in self.config["power_keywords"]):
            if dec is not None:
                power_midspan_heights.append((dec, fmt))

    def _assign_comm_attachments(self, telecom_data, comm_dict, mapped_comms):
        """Sort and assign telecom attachments to comm1-4, and capture ALL communication heights"""
        logging.debug(f"Raw telecom data before filtering: {telecom_data}")
        logging.debug(f"Mapped comms: {mapped_comms}")
        logging.debug(f"Comm dict keys: {list(comm_dict.keys())}")
        
        # Filter out entries without measured data
        filtered_telecom_data = [x for x in telecom_data if x[2]]
        
        # Updated filter to include expanded communication keywords
        # Include: 'CATV Com', 'Telco Com', 'Fiber Optic Com', 'insulator', 'Power Guy'
        keywords = self.config.get("comm_keywords", ['catv com', 'telco com', 'fiber optic com', 'insulator', 'power guy', 'catv', 'telco', 'fiber', 'communication', 'comm'])
        keyword_filtered = [x for x in filtered_telecom_data if any(kw in str(x[2]).lower() for kw in keywords)]
        
        # If keyword matches are present, use them; otherwise, use the full filtered list
        # This ensures we don't accidentally exclude valid communication attachments
        if keyword_filtered:
            filtered_telecom_data = keyword_filtered
            logging.debug(f"Using keyword-filtered data: {len(keyword_filtered)} entries")
        else:
            logging.debug(f"No keyword matches found, using all filtered data: {len(filtered_telecom_data)} entries")
            # Log what we're including when no keyword matches are found
            if filtered_telecom_data:
                logging.debug(f"Non-keyword filtered entries: {[x[2] for x in filtered_telecom_data[:5]]}")  # Log first 5
        
        # Sort data (highest to lowest)
        filtered_telecom_data.sort(key=lambda x: x[0], reverse=True)
        logging.debug(f"Sorted telecom data for comm assignment: {filtered_telecom_data}")
        
        # Determine if this is attachment data or midspan data based on comm_dict keys
        is_midspan_data = any(key.endswith('_Midspan') for key in comm_dict.keys())
        
        # Assign to comm fields (first 4 only)
        comm_names = ["comm1", "comm2", "comm3", "comm4"]
        for i, comm in enumerate(comm_names):
            if is_midspan_data:
                # For midspan data, always use the _Midspan suffix
                key = f"{comm}_Midspan"
            else:
                # For attachment data, use the base comm name (comm1, comm2, etc.)
                key = comm
            
            logging.debug(f"Checking {comm}: key={key}, key in comm_dict={key in comm_dict}, is_midspan_data={is_midspan_data}")
            if key in comm_dict and i < len(filtered_telecom_data):
                comm_dict[key] = filtered_telecom_data[i][1]
                logging.info(f"Assigned {filtered_telecom_data[i][1]} to {key}")
            else:
                logging.debug(f"Skipping assignment for {comm}: key={key}, key in comm_dict={key in comm_dict}, i={i}, data_len={len(filtered_telecom_data)}")
        
        # NEW: Create comprehensive summary of ALL communication attachment heights (only for attachment data, not midspan)
        if filtered_telecom_data and not is_midspan_data:
            all_comm_heights = []
            for height_decimal, height_formatted, provider in filtered_telecom_data:
                # Include provider info if available
                if provider and str(provider).strip():
                    all_comm_heights.append(f"{height_formatted} ({provider})")
                else:
                    all_comm_heights.append(height_formatted)
            
            # Add comprehensive field to comm_dict
            comm_dict['All_Comm_Heights'] = '; '.join(all_comm_heights)
            logging.info(f"All communication heights captured: {comm_dict['All_Comm_Heights']}")
            
            # Also add count of total communication attachments
            comm_dict['Total_Comm_Count'] = str(len(filtered_telecom_data))
            logging.info(f"Total communication attachments found: {len(filtered_telecom_data)}")
        else:
            logging.debug("No filtered telecom data found for assignment or this is midspan data")
    
    def _calculate_power_heights(self, power_heights, power_midspan_heights, telecom_heights):
        """Calculate lowest power heights"""
        lowest_power = ""
        lowest_power_midspan = ""
        
        if power_heights:
            min_threshold = max(telecom_heights) if telecom_heights else 0
            valid_power = [(h, f) for h, f in power_heights if h >= min_threshold]
            
            if valid_power:
                lowest_power = min(valid_power, key=lambda x: x[0])[1]
            else:
                lowest_power = min(power_heights, key=lambda x: x[0])[1]
        
        if power_midspan_heights:
            lowest_power_midspan = min(power_midspan_heights, key=lambda x: x[0])[1]
        
        return {
            'Power Height': lowest_power,
            'Power Midspan': lowest_power_midspan
        }
    
    def _match_metronet(self, owner):
        """Check if owner matches Proposed MetroNet (case insensitive) 
           For 'power guy' keyword, still requires company/owner name to be present."""
        owner_str = str(owner).lower()
        keywords = self.config["telecom_keywords"].get("Proposed MetroNet", [])
        
        # Check regular MetroNet keywords first
        for keyword in keywords:
            if keyword.lower() in owner_str:
                return True
        
        # Special handling for 'power guy' - must have company name present
        if "power guy" in owner_str:
            # Check if any company/provider names are also present
            all_providers = set()
            for provider_keywords in self.config["telecom_keywords"].values():
                all_providers.update([k.lower() for k in provider_keywords])
            
            # Also check power company names
            power_keywords = [k.lower() for k in self.config.get("power_keywords", [])]
            all_providers.update(power_keywords)
            
            # If any provider/company name is found along with 'power guy', it's valid
            for provider in all_providers:
                if provider in owner_str and provider != "power guy":
                    return True
        
        return False
    
    def _match_telecom_provider(self, owner):
        """Match owner to telecom provider (case insensitive)"""
        owner_str = str(owner).lower()
        for provider, keywords in self.config["telecom_keywords"].items():
            if any(k.lower() in owner_str for k in keywords):
                return provider
        return None
    
    def _get_pole_address(self, node):
        """Get pole address from geocoding cache or service"""
        lat, lon = node.get('latitude'), node.get('longitude')
        if not lat or not lon:
            return ''
            
        if self.geocoder:
            # If geocoding is enabled, try cache first, then geocoding service
            address = self.geocoder.reverse(lat, lon)
            return address
        else:
            # If geocoding is disabled, only check cache
            try:
                key = f"{round(float(lat), 7)},{round(float(lon), 7)}"
                return self.geocoder.cache.get(key, '')
            except:
                return ''
    
    def _format_span_distance(self, span_distance):
        """Format span distance as rounded feet with ' suffix"""
        if not span_distance:
            return ""
        
        try:
            # Convert to float and round to nearest whole number
            distance_feet = round(float(span_distance))
            return f"{distance_feet}'"
        except (ValueError, TypeError):
            # If conversion fails, return original value
            return str(span_distance)

    def _format_pole_height_class(self, node):
        """Format pole height and class"""
        pole_height = node.get('pole_height', '')
        pole_class = node.get('pole_class', '')
        
        if pole_height and pole_class:
            try:
                return f"{int(float(pole_height))}>{int(float(pole_class))}"
            except:
                return f"{pole_height}>{pole_class}"
        return ""

    def _count_existing_risers(self, node):
        """Count existing risers excluding MetroNet"""
        count = 0
        for key, val in node.items():
            if key.startswith("POA_") and not key.endswith("HT"):
                if isinstance(val, str):
                    owner = val.lower()
                    if "riser" in owner and not self._match_metronet(val):
                        count += 1
        return str(count)
    
    def _get_tension_mappings(self):
        """Get tension field mappings from column mappings configuration"""
        tension_mappings = {}
        
        for element, attribute, output in self.mapping_data:
            if attribute == "Tension":
                tension_mappings[element] = output
        
        return tension_mappings
    
    def _parse_span_length(self, span_distance):
        """Parse span length from span distance string"""
        try:
            if not span_distance:
                return None
                
            # Convert to string and clean
            span_str = str(span_distance).strip()
            if not span_str or span_str.lower() in ['nan', 'none', '']:
                return None
            
            # Remove common suffixes like ' or ft
            span_str = span_str.replace("'", "").replace("ft", "").replace("feet", "").strip()
            
            # Try to convert to float
            return float(span_str)
            
        except (ValueError, TypeError):
            logging.warning(f"Could not parse span length: {span_distance}")
            return None
    
    def _parse_height_value(self, value):
        """Parse height value from various formats including feet'inches" """
        try:
            # Handle NaN/None values early
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return None

            # Use shared utility to parse into decimal-feet
            decimal_feet = Utils.parse_height_decimal(value)

            if decimal_feet is not None:
                # Round to two decimals for consistency
                return round(decimal_feet, 2)

            logging.debug(f"Utils.parse_height_decimal was unable to parse '{value}' – attempting direct float conversion")

            # Fallback: strip quotes/units and try float conversion
            value_str = str(value).replace("'", "").replace('"', "").replace("ft", "").replace("feet", "").strip()
            return round(float(value_str), 2)

        except (ValueError, TypeError) as e:
            logging.warning(f"Could not parse height value '{value}': {e}")
            return None
    
    def write_output(self, result_data, output_file):
        """Write processed data to Excel output file"""
        try:
            if not result_data:
                logging.warning("No data to write to output file")
                return

            # Filter out None or non-dict items before sorting
            filtered_data = [item for item in result_data if item and isinstance(item, dict)]
            if not filtered_data:
                logging.warning("No valid data to write after filtering")
                return

            # Sort data only if QC file is not active (preserve QC order when active)
            if self.qc_reader and self.qc_reader.is_active():
                # QC file is active - preserve the exact order from QC file
                sorted_data = filtered_data
                logging.info("QC file is active - preserving QC order for main sheet")
            else:
                # No QC file - sort by pole SCID as usual
                sorted_data = sorted(filtered_data, key=lambda x: Utils.extract_numeric_part(x.get('Pole', '')))
                logging.info("No QC file - sorting main sheet by pole SCID")

            # Create data cache for QC sheet population
            self._processed_data_cache = {}
            for row in sorted_data:
                pole = row.get('Pole', '').strip()
                if pole:
                    self._processed_data_cache[pole] = row

            # Check if output file exists, if not, try to create it from template
            from pathlib import Path
            output_path = Path(output_file)
            
            if not output_path.exists():
                # Try to find and copy template
                template_path = None
                if hasattr(self, 'config') and self.config:
                    # Look for template in common locations
                    possible_templates = [
                        'C:/Users/nsaro/Desktop/Test/Consumer SS Template.xltm',
                        'Consumer SS Template.xltm',
                        'template.xlsx',
                        'template.xltm'
                    ]
                    
                    for template in possible_templates:
                        if Path(template).exists():
                            template_path = template
                            break
                
                if template_path:
                    logging.info(f"Output file doesn't exist, copying from template: {template_path}")
                    import shutil
                    try:
                        shutil.copy2(template_path, output_file)
                        logging.info(f"Successfully created output file from template")
                    except Exception as e:
                        logging.error(f"Failed to copy template: {e}")
                        return
                else:
                    logging.error(f"Output file '{output_file}' doesn't exist and no template found")
                    return

            # Validate the output file after creation/copying
            if not output_path.exists() or output_path.stat().st_size == 0:
                logging.error(f"Output file '{output_file}' is missing or empty.")
                return

            # Attempt to load the workbook inside a try/except block to catch EOFError
            try:
                # Use keep_vba=True to preserve macros in .xlsm files
                wb = load_workbook(output_file, keep_vba=True)
            except EOFError as eof_error:
                logging.error(f"EOFError encountered when loading workbook '{output_file}': {eof_error}. The template file may be corrupted.")
                return

            # Determine worksheet to use
            if hasattr(self, 'config') and self.config:
                worksheet_name = self.config.get('output_settings', {}).get('worksheet_name', 'Consumers pg1')
            else:
                worksheet_name = 'Consumers pg1'

            if worksheet_name in wb.sheetnames:
                ws = wb[worksheet_name]
            else:
                ws = wb.active
                logging.warning(f"Worksheet '{worksheet_name}' not found, using '{ws.title}'")

            # Apply final span length tolerance check to main sheet data
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("Applying final span length tolerance check to main sheet data")
                tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                logging.info(f"Span length tolerance setting: {tolerance}")
                
                tolerance_updates = 0
                for row_data in sorted_data:
                    pole = row_data.get('Pole', '')
                    to_pole = row_data.get('To Pole', '')
                    
                    if pole and to_pole:
                        # Get QC span length
                        qc_span = self.qc_reader.get_qc_span_length(pole, to_pole)
                        excel_span = row_data.get('Span Length', '')
                        
                        # Log every connection for debugging
                        if qc_span or excel_span:
                            logging.info(f"Tolerance check: {pole} -> {to_pole}: Excel='{excel_span}', QC='{qc_span}'")
                        
                        if qc_span and excel_span:
                            # Apply tolerance check
                            final_span = self._apply_span_length_tolerance(excel_span, qc_span, tolerance)
                            
                            # Always update with the final span length (tolerance method handles the logic)
                            row_data['Span Length'] = final_span
                            
                            # Log the result for debugging
                            if final_span != excel_span:
                                logging.info(f"Final tolerance update: {pole} -> {to_pole}: '{excel_span}' -> '{final_span}' (QC: {qc_span}, tolerance: {tolerance})")
                                tolerance_updates += 1
                            else:
                                logging.info(f"Excel span length retained: {pole} -> {to_pole}: '{excel_span}' (QC: {qc_span}, tolerance: {tolerance})")
                            
                
                logging.info(f"Completed span length tolerance check: {tolerance_updates} updates applied")
            
            # Write data; using mapped writing if available, else a simple write
            if hasattr(self, 'mapping_data') and self.mapping_data:
                self._write_data_to_worksheet(ws, sorted_data, self.mapping_data)
            else:
                self._write_data_simple(ws, sorted_data)

            # Automatically populate QC sheet if QC reader is active
            if self.qc_reader and self.qc_reader.is_active():
                logging.info("QC reader is active - populating QC sheet")
                self._populate_qc_sheet(wb)
                
                # Add conditional formatting to compare main sheet and QC sheet
                logging.info("Adding conditional formatting to compare main sheet and QC sheet")
                self._add_sheet_comparison_formatting(wb, worksheet_name)
            else:
                logging.info("QC reader not active - skipping QC sheet population")

            wb.save(output_file)
            logging.info(f"Successfully wrote {len(sorted_data)} records to {output_file}")

        except Exception as e:
            logging.error(f"Error writing output: {e}")
            raise

    def _populate_qc_sheet(self, workbook):
        """Automatically populate QC sheet with data from QC file into existing columns"""
        try:
            import re
            from openpyxl.utils import get_column_letter
            
            # Check if QC sheet exists
            if "QC" not in workbook.sheetnames:
                logging.info("No existing QC sheet found, skipping QC data population")
                return
                
            qc_sheet = workbook["QC"]
            logging.info("Found existing QC sheet, populating with QC data")
            
            # Get complete row data from QC file
            qc_data_rows = self.qc_reader.get_qc_data_rows()
            
            if not qc_data_rows:
                logging.warning("No QC data rows found to populate")
                return
            
            logging.info(f"QC data rows to populate: {len(qc_data_rows)}")
            
            # Find the header row in the existing QC sheet (try rows 1, 2, 3)
            header_row = None
            existing_headers = {}
            
            for row_num in [1, 2, 3]:
                headers_found = {}
                for col_idx in range(1, min(qc_sheet.max_column + 1, 50)):  # Limit column scan to 50 columns
                    cell_value = qc_sheet.cell(row=row_num, column=col_idx).value
                    if cell_value:
                        # Clean header text (remove extra spaces, newlines)
                        header_text = re.sub(r"\s+", " ", str(cell_value).replace("\n", " ")).strip()
                        if header_text:
                            headers_found[header_text] = col_idx
                
                # Check if this row has the required columns
                if any(header in headers_found for header in ['Pole', 'To Pole']):
                    header_row = row_num
                    existing_headers = headers_found
                    logging.info(f"Found headers in row {header_row}: {list(existing_headers.keys())}")
                    break
            
            if not header_row:
                logging.warning("Could not find header row with 'Pole' and 'To Pole' columns in QC sheet")
                return
            
            # Clear existing data rows (keep headers intact) - but only clear what we need
            data_start_row = header_row + 1
            max_cols_to_clear = min(qc_sheet.max_column, 50)  # Limit to 50 columns
            
            # Find the actual end of data to avoid clearing unnecessary rows
            actual_max_row = data_start_row
            for row_idx in range(data_start_row, min(qc_sheet.max_row + 1, data_start_row + 10000)):  # Limit to 10,000 rows
                has_data = False
                for col_idx in range(1, max_cols_to_clear + 1):
                    if qc_sheet.cell(row=row_idx, column=col_idx).value:
                        has_data = True
                        break
                if has_data:
                    actual_max_row = row_idx
                else:
                    break  # Stop at first empty row
            
            # Only clear rows that actually have data
            rows_to_clear = min(actual_max_row - data_start_row + 1, len(qc_data_rows) + 100)  # Clear existing data + some buffer
            logging.info(f"Clearing {rows_to_clear} rows starting from row {data_start_row}")
            
            if rows_to_clear > 0:
                for row_idx in range(data_start_row, data_start_row + rows_to_clear):
                    for col_idx in range(1, max_cols_to_clear + 1):
                        qc_sheet.cell(row=row_idx, column=col_idx).value = None
            
            # Create mapping from QC file columns to existing sheet columns
            column_mapping = {}
            
            # Log available columns for debugging
            qc_columns = list(qc_data_rows[0].keys()) if qc_data_rows else []
            logging.info(f"QC file columns: {qc_columns}")
            logging.info(f"QC sheet headers: {list(existing_headers.keys())}")
            
            # Enhanced column mapping with fuzzy matching
            for qc_column in qc_columns:
                mapped = False
                
                # Try exact match first
                if qc_column in existing_headers:
                    column_mapping[qc_column] = existing_headers[qc_column]
                    mapped = True
                else:
                    # Try case-insensitive match
                    for existing_header, col_idx in existing_headers.items():
                        if qc_column.lower() == existing_header.lower():
                            column_mapping[qc_column] = col_idx
                            mapped = True
                            break
                
                # Try normalized matching (ignoring spaces, punctuation, newlines)
                if not mapped:
                    for existing_header, col_idx in existing_headers.items():
                        if self._columns_match(qc_column, existing_header):
                            column_mapping[qc_column] = col_idx
                            logging.info(f"Normalized match '{qc_column}' -> '{existing_header}'")
                            mapped = True
                            break
                
                if not mapped:
                    logging.warning(f"Could not map QC column '{qc_column}' to any QC sheet column")
            
            logging.info(f"Column mapping: {column_mapping}")
            logging.info(f"Mapped {len(column_mapping)} out of {len(qc_columns)} QC file columns")
            
            # Populate data rows
            rows_written = 0
            for row_idx, row_data in enumerate(qc_data_rows):
                sheet_row = data_start_row + row_idx
                
                # Populate mapped columns from QC file
                for qc_column, value in row_data.items():
                    if qc_column in column_mapping:
                        col_idx = column_mapping[qc_column]
                        qc_sheet.cell(row=sheet_row, column=col_idx).value = value
                
                # Try to populate missing columns from main data if available
                self._populate_missing_qc_columns(qc_sheet, sheet_row, row_data, existing_headers, column_mapping)
                
                rows_written += 1
                
                # Progress logging for large datasets
                if rows_written % 100 == 0:
                    logging.info(f"Populated {rows_written} QC rows...")
            
            logging.info(f"Successfully populated QC sheet with {rows_written} rows into {len(column_mapping)} matching columns")
            
        except Exception as e:
            logging.error(f"Error populating QC sheet: {e}")
            # Don't raise the exception - QC sheet population is optional
    
    def _columns_match(self, qc_column, sheet_column):
        """Check if QC file column matches QC sheet column by normalizing spaces, punctuation, and newlines"""
        import re
        
        def normalize_column_name(name):
            # Remove newlines and replace with spaces
            name = name.replace('\n', ' ').replace('\r', ' ')
            # Remove all punctuation and special characters, keep only letters, numbers, and spaces
            name = re.sub(r'[^\w\s]', ' ', name)
            # Replace multiple spaces with single space and strip
            name = re.sub(r'\s+', ' ', name).strip()
            return name.lower()
        
        qc_norm = normalize_column_name(qc_column)
        sheet_norm = normalize_column_name(sheet_column)
        
        # Direct match after normalization
        return qc_norm == sheet_norm
    
    def _populate_missing_qc_columns(self, qc_sheet, sheet_row, row_data, existing_headers, column_mapping):
        """Populate missing QC columns with data from main processing if available"""
        try:
            # Get pole and to_pole from current row
            pole = row_data.get('Pole', '').strip()
            to_pole = row_data.get('To Pole', '').strip()
            
            if not pole:
                return
            
            # Check if we have processed data for this pole
            if hasattr(self, '_processed_data_cache'):
                pole_data = self._processed_data_cache.get(pole)
                if pole_data:
                    # Populate Pole Address if missing
                    if 'Pole Address (if available)' in existing_headers and 'Pole Address (if available)' not in column_mapping:
                        address = pole_data.get('Pole Address', '')
                        if address:
                            col_idx = existing_headers['Pole Address (if available)']
                            qc_sheet.cell(row=sheet_row, column=col_idx).value = address
                    
                    # Populate Proposed height if missing
                    if 'Proposed height of new attachment point' in existing_headers and 'Proposed height of new attachment point' not in column_mapping:
                        proposed_height = pole_data.get('Proposed MetroNet', '')
                        if proposed_height:
                            col_idx = existing_headers['Proposed height of new attachment point']
                            qc_sheet.cell(row=sheet_row, column=col_idx).value = proposed_height
                    
                    # Populate other missing columns as needed
                    missing_mappings = {
                        'Secondary or Neutral Power Height (Height of Lowest Power Conductor or Equipment, excluding streetlights)': 'Power Height',
                        'Pole Height & Class': 'Pole Height & Class',
                        'Pole to Pole Span Length (from starting point)': 'Span Length',
                        'Final Mid Span Ground Clearance of Proposed Attachment': 'Proposed MetroNet_Midspan',
                        'Guy Size': 'Guy Size',
                        'Guy Lead': 'Guy Lead',
                        'Guy Direction': 'Guy Direction',
                        'Notes (Items that need to be performed by Consumers Energy or other Companies)': 'Notes'
                    }
                    
                    for qc_header, data_key in missing_mappings.items():
                        if qc_header in existing_headers and qc_header not in [list(existing_headers.keys())[col_idx-1] for col_idx in column_mapping.values()]:
                            value = pole_data.get(data_key, '')
                            if value:
                                col_idx = existing_headers[qc_header]
                                qc_sheet.cell(row=sheet_row, column=col_idx).value = value
        
        except Exception as e:
            logging.debug(f"Error populating missing QC columns for row {sheet_row}: {e}")
    
    def _get_internal_key(self, element, attribute):
        """Get internal key for mapping"""
        mappings = {
            "Pole": {
                "Number": "Pole",
                "Map": "Map",
                "Address": "Address", 
                "Height & Class": "Pole Height/Class",
                "MR Notes": "Notes",
                "To Pole": "To Pole",
                "Latitude": "Latitude",
                "Longitude": "Longitude",
                "Tag": "Pole Tag",
                "Number of Existing Risers": "Existing Risers"
            },
            "New Guy": {
                "Size": "Guy Size",
                "Lead": "Guy Lead", 
                "Direction": "Guy Direction"
            },
            "Power": {
                "Lowest Height": "Power Height",
                "Lowest Midspan": "Power Midspan"
            },
            "Span": {
                "Length": "Span Length"
            },
            "System": {
                "Line Number": "Line No."
            },
            "Street Light": {
                "Lowest Height": "Street Light Height"
            },
            "Cable": {
                "Tension": "Cable Tension",
                "Type1": "Cable Type 1",
                "Diameter1": "Cable Diameter 1",
                "Type2": "Cable Type 2", 
                "Diameter2": "Cable Diameter 2",
                "Total Bundle Diameter": "Total Bundle Diameter"
            }
        }
        
        if element in mappings:
            return mappings[element].get(attribute)
        elif element in ["comm1", "comm2", "comm3", "comm4"]:
            if attribute == "Attachment Ht":
                return element
            elif attribute == "Midspan Ht":
                return f"{element}_Midspan"
            elif attribute == "Tension":
                return f"{element}_Tension"
        elif element in self.config["telecom_providers"]:
            if attribute == "Attachment Ht":
                return element
            elif attribute == "Midspan Ht":
                # Special case for Proposed MetroNet midspan
                if element == "Proposed MetroNet":
                    return "Proposed MetroNet_Midspan"
                else:
                    return f"{element}_Midspan"
            elif attribute == "Tension":
                # Special case for Proposed MetroNet tension
                if element == "Proposed MetroNet":
                    return "Heavy Loaded Tension (NESC Rule 251)"
                else:
                    return f"{element}_Tension"
        
        return None

    def generate_output_file(self, job_name, template_path):
        """Generate output file by copying template with job name, preserving file extension."""
        import shutil
        from pathlib import Path
        
        template = Path(template_path)
        if not template.exists():
            logging.error(f"Template file not found: {template_path}")
            return None
        
        # Preserve the original file extension (.xlsx or .xlsm)
        template_extension = template.suffix
        output_file = template.parent / f"{job_name} Spread Sheet{template_extension}"
        
        try:
            shutil.copy2(template, output_file)
            
            # Verify the copy was successful
            if not output_file.exists() or output_file.stat().st_size == 0:
                logging.error(f"Copied output file '{output_file}' is empty. Check the template file.")
                return None
            logging.info(f"Successfully copied template to: {output_file}")
            return output_file
        except Exception as e:
            logging.error(f"Error copying template file: {e}")
            return None
    
    def _write_data_to_worksheet(self, ws, sorted_data, mapping_data):
        """Write sorted_data to worksheet ws using mapping_data for column mapping."""
        import re
        # Get config settings
        header_row = self.config.get("output_settings", {}).get("header_row", 1)
        data_start_row = self.config.get("output_settings", {}).get("data_start_row", header_row + 2)

        # Get headers from the worksheet
        headers = []
        for cell_obj in ws[header_row]:
            if cell_obj.value:
                header_text = re.sub(r"\s+", " ", str(cell_obj.value).replace("\n", " ")).strip()
                headers.append(header_text)
            else:
                headers.append("")
        col_map = {h: idx + 1 for idx, h in enumerate(headers) if h.strip()}
        logging.info(f"Found worksheet headers: {list(col_map.keys())}")

        # Build mapping from internal key to Excel column name
        internal_to_excel = {}
        for element, attribute, output_col_name in mapping_data:
            internal_key = self._get_internal_key(element, attribute)
            if internal_key and output_col_name.strip():
                internal_to_excel[internal_key] = output_col_name
                logging.debug(f"Mapping {element}:{attribute} -> {internal_key} -> {output_col_name}")
        
        logging.info(f"Internal to Excel mappings: {internal_to_excel}")
        
        successful_writes = 0
        missing_columns = set()
        
        for i, data_row_content in enumerate(sorted_data, start=1):
            data_row_content['Line No.'] = i
            
            # Log tension-related fields for debugging
            tension_fields = {k: v for k, v in data_row_content.items() if 'tension' in k.lower()}
            if tension_fields:
                logging.info(f"Row {i} tension fields: {tension_fields}")
            
            for internal_name, value in data_row_content.items():
                excel_col_name = internal_to_excel.get(internal_name, "")
                if not excel_col_name:
                    if 'tension' in internal_name.lower():
                        logging.warning(f"No Excel column mapping found for tension field: {internal_name}")
                    continue
                
                col = col_map.get(excel_col_name)
                if col:
                    try:
                        cell_to_write = ws.cell(row=data_start_row + i - 1, column=col)
                        cell_to_write.value = value
                        if 'tension' in internal_name.lower():
                            logging.info(f"Writing tension value {value} to column {excel_col_name} (internal: {internal_name})")
                        successful_writes += 1
                    except Exception as e:
                        logging.warning(f"Error writing cell: {e}")
                else:
                    missing_columns.add(excel_col_name)
                    if 'tension' in excel_col_name.lower():
                        logging.warning(f"Column not found in worksheet: {excel_col_name} (internal: {internal_name})")
        if missing_columns:
            logging.info(f"Note: Some mapped columns not found in template: {', '.join(sorted(missing_columns))}")
        else:
            logging.info("All mapped columns found in template")
        logging.info(f"Successfully wrote {successful_writes} data cells")
    
    def _write_data_simple(self, ws, sorted_data):
        """Fallback: Write sorted_data to worksheet ws with no mapping (just as columns in order)."""
        # Note: Removed conditional formatting as requested
        
        for row_idx, row_data in enumerate(sorted_data, start=1):
            # Check if this row represents a QC mismatch (for logging only)
            is_qc_mismatch = False
            if self.qc_reader and self.qc_reader.is_active():
                pole = row_data.get('Pole', '')
                to_pole = row_data.get('To Pole', '')
                if pole and to_pole:
                    is_qc_mismatch = not self.qc_reader.has_connection(pole, to_pole)
                    if is_qc_mismatch:
                        logging.debug(f"QC mismatch detected for row {row_idx}: {pole} -> {to_pole}")
            
            for col_idx, value in enumerate(row_data.values(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Note: Removed conditional formatting/highlighting as requested

    def _process_qc_filtered_connections(self, connections_df, mappings, sections_df):
        """Process connections when QC file is active - use EXACT QC Pole and ToPole values in specified order"""
        result_data = []
        
        # Get ordered connections from QC file in ORIGINAL format
        qc_original_connections = self.qc_reader.get_original_ordered_connections()
        qc_normalized_connections = self.qc_reader.get_ordered_connections()
        
        logging.info(f"Processing {len(qc_original_connections)} QC connections in specified order")
        logging.info("QC Mode: Using EXACT original Pole and ToPole format from QC file")
        
        # Create lookup for connection data from Excel (bidirectional)
        connection_lookup = {}
        for _, conn in connections_df.iterrows():
            n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
            if n1 in mappings['node_id_to_scid'] and n2 in mappings['node_id_to_scid']:
                scid1 = mappings['node_id_to_scid'][n1]
                scid2 = mappings['node_id_to_scid'][n2]
                
                conn_info = {
                    'connection_id': conn.get('connection_id', ''),
                    'span_distance': conn.get('span_distance', ''),
                    'node1_id': n1,
                    'node2_id': n2
                }
                
                # Store connection lookup (use sorted tuple as key to avoid duplication)
                connection_key = tuple(sorted([scid1, scid2]))
                connection_lookup[connection_key] = conn_info
        
        # Process QC connections in the exact order specified in QC file
        for i, (qc_pole_orig, qc_to_pole_orig) in enumerate(qc_original_connections):
            # Get the corresponding normalized versions for data lookup
            qc_pole_norm, qc_to_pole_norm = qc_normalized_connections[i]
            # Check if this connection exists in Excel data using normalized SCIDs
            connection_key = tuple(sorted([qc_pole_norm, qc_to_pole_norm]))
            conn_info = connection_lookup.get(connection_key)
            
            if not conn_info:
                logging.warning(f"QC connection {qc_pole_orig} -> {qc_to_pole_orig} not found in Excel data")
                # Always create a row for QC connections, even if no data is available
                pole_node_data = mappings['scid_to_row'].get(qc_pole_norm, {})
                to_pole_node_data = mappings['scid_to_row'].get(qc_to_pole_norm, {})
                
                # Try to find span distance from connections_df using different lookup approaches
                span_distance = ''
                
                # Try direct SCID lookup in connections_df
                for _, conn in connections_df.iterrows():
                    n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
                    
                    # Check if either node matches our SCIDs (direct or through mapping)
                    scid1 = mappings['node_id_to_scid'].get(n1, n1)
                    scid2 = mappings['node_id_to_scid'].get(n2, n2)
                    
                    # Check all possible combinations
                    if ((scid1 == qc_pole_norm and scid2 == qc_to_pole_norm) or
                        (scid1 == qc_to_pole_norm and scid2 == qc_pole_norm) or
                        (n1 == qc_pole_norm and n2 == qc_to_pole_norm) or
                        (n1 == qc_to_pole_norm and n2 == qc_pole_norm)):
                        span_distance = conn.get('span_distance', '')
                        if span_distance:
                            logging.info(f"Found span distance {span_distance} for QC connection {qc_pole_orig} -> {qc_to_pole_orig}")
                            break
                
                # If no exact match, try alternative SCID matching (e.g., "118 MISM013" -> "118")
                if not span_distance:
                    # Extract base SCID numbers for alternative matching
                    import re
                    
                    def extract_base_scid(scid):
                        """Extract base SCID number (e.g., '118 MISM013' -> '118')"""
                        match = re.match(r'^(\d+)', str(scid).strip())
                        return match.group(1) if match else scid
                    
                    qc_pole_base = extract_base_scid(qc_pole_norm)
                    qc_to_pole_base = extract_base_scid(qc_to_pole_norm)
                    
                    for _, conn in connections_df.iterrows():
                        n1, n2 = str(conn['node_id_1']).strip(), str(conn['node_id_2']).strip()
                        scid1 = mappings['node_id_to_scid'].get(n1, n1)
                        scid2 = mappings['node_id_to_scid'].get(n2, n2)
                        
                        scid1_base = extract_base_scid(scid1)
                        scid2_base = extract_base_scid(scid2)
                        
                        # Check base SCID matching
                        if ((scid1_base == qc_pole_base and scid2_base == qc_to_pole_base) or
                            (scid1_base == qc_to_pole_base and scid2_base == qc_pole_base)):
                            span_distance = conn.get('span_distance', '')
                            if span_distance:
                                logging.info(f"Found span distance {span_distance} for QC connection {qc_pole_orig} -> {qc_to_pole_orig} using base SCID matching ({scid1} <-> {scid2})")
                                break
                
                # Create connection info with found span distance (or empty if not found)
                conn_info = {
                    'connection_id': '',
                    'span_distance': span_distance,
                    'node1_id': '',
                    'node2_id': ''
                }
            
            # Get node data for the pole specified in QC file (using normalized SCID for lookup)
            pole_node_data = mappings['scid_to_row'].get(qc_pole_norm, {})
            
            # Always create output row for QC connections, even if no data is available
            # This ensures both main sheet and QC sheet have exactly the same connections
            row_data = self._create_qc_output_row(
                qc_pole_orig,  # Use ORIGINAL format from QC file
                qc_to_pole_orig,  # Use ORIGINAL format from QC file
                qc_pole_norm,  # Pass normalized version for data lookup
                qc_to_pole_norm,  # Pass normalized version for data lookup
                conn_info, 
                pole_node_data, 
                mappings['scid_to_row'], 
                sections_df
            )
            
            if row_data:
                result_data.append(row_data)
                logging.debug(f"Added QC connection (exact original): {qc_pole_orig} -> {qc_to_pole_orig}")
            else:
                # If _create_qc_output_row fails, create a minimal row to ensure connection is included
                logging.info(f"Creating minimal row for QC connection: {qc_pole_orig} -> {qc_to_pole_orig}")
                minimal_row = {
                    'Pole': qc_pole_orig,
                    'To Pole': qc_to_pole_orig,
                    'Line No.': '',
                    'Span Length': conn_info.get('span_distance', ''),
                    'Pole Height & Class': '',
                    'Power Height': '',
                    'Streetlight (bottom of bracket)': '',
                    'Guy Size': '',
                    'Guy Lead': '',
                    'Guy Direction': '',
                    'Pole Address': '',
                    'Notes': 'QC connection - limited data available'
                }
                result_data.append(minimal_row)
                logging.debug(f"Added minimal QC connection: {qc_pole_orig} -> {qc_to_pole_orig}")
        
        logging.info(f"Generated {len(result_data)} QC-filtered output rows in exact QC order")
        return result_data
    
    def _create_qc_output_row(self, pole_orig, to_pole_orig, pole_norm, to_pole_norm, conn_info, pole_node_data, scid_to_row, sections_df):
        """Create output row for QC filtering using exact ORIGINAL QC Pole and ToPole values"""
        # Try to create a row using the standard method first
        row_data = self._create_output_row(pole_norm, to_pole_norm, conn_info, pole_node_data, scid_to_row, sections_df)
        
        if row_data:
            # Force the exact ORIGINAL QC values (override any logic that might change them)
            row_data['Pole'] = pole_orig
            row_data['To Pole'] = to_pole_orig
            
            # Apply span length tolerance logic if QC reader is available
            if self.qc_reader and self.qc_reader.is_active():
                logging.info(f"Checking span length tolerance for {pole_orig} -> {to_pole_orig}")
                qc_span = self.qc_reader.get_qc_span_length(pole_orig, to_pole_orig)
                logging.info(f"QC span for {pole_orig} -> {to_pole_orig}: '{qc_span}'")
                
                if qc_span:
                    excel_span = row_data.get('Span Length', '')
                    tolerance = self.config.get('processing_options', {}).get('span_length_tolerance', 3)
                    logging.info(f"Excel span: '{excel_span}', QC span: '{qc_span}', tolerance: {tolerance}")
                    
                    # Apply tolerance check and use QC span if within tolerance
                    final_span = self._apply_span_length_tolerance(excel_span, qc_span, tolerance)
                    row_data['Span Length'] = final_span
                    logging.info(f"Final span length for {pole_orig} -> {to_pole_orig}: '{final_span}'")
                else:
                    logging.info(f"No QC span length found for {pole_orig} -> {to_pole_orig}")
            
            logging.debug(f"QC Row: Pole={pole_orig}, To Pole={to_pole_orig} (original format preserved)")
        else:
            # If standard method fails, create a minimal row to ensure QC connection is included
            logging.debug(f"Creating minimal QC row for {pole_orig} -> {to_pole_orig}")
            row_data = {
                'Pole': pole_orig,
                'To Pole': to_pole_orig,
                'Line No.': '',
                'Span Length': self._format_span_distance(conn_info.get('span_distance', '')),
                'Pole Height & Class': '',
                'Power Height': '',
                'Streetlight (bottom of bracket)': '',
                'Guy Size': '',
                'Guy Lead': '',
                'Guy Direction': '',
                'Pole Address': '',
                'Notes': 'QC connection - limited data available',
                # Add empty values for communication fields
                'comm1': '',
                'comm2': '',
                'comm3': '',
                'comm4': '',
                'Proposed MetroNet': '',
                'Verizon': '',
                'AT&T': '',
                'Comcast': '',
                'Zayo': '',
                'Jackson ISD': '',
                'All_Comm_Heights': '',
                'Total_Comm_Count': '',
                'Power Midspan': '',
                'Street Light Height': '',
                'Existing Risers': '',
                'Map': ''
            }
        
        return row_data

    def _add_sheet_comparison_formatting(self, workbook, main_sheet_name):
        """Conditional formatting has been disabled as requested - logging comparison info instead"""
        try:
            # Check if both sheets exist
            if main_sheet_name not in workbook.sheetnames or "QC" not in workbook.sheetnames:
                logging.info("Cannot compare sheets - missing main sheet or QC sheet")
                return
            
            main_sheet = workbook[main_sheet_name]
            qc_sheet = workbook["QC"]
            
            # Get config settings for data range
            header_row = self.config.get("output_settings", {}).get("header_row", 3)
            data_start_row = self.config.get("output_settings", {}).get("data_start_row", 4)
            
            # Find the data range in main sheet
            max_row = main_sheet.max_row
            max_col = min(main_sheet.max_column, 50)  # Limit to 50 columns for performance
            
            if max_row < data_start_row:
                logging.info("No data rows found in main sheet for comparison")
                return
            
            logging.info(f"Sheet comparison available between {main_sheet_name} and QC sheets")
            logging.info(f"Data range: {max_col} columns from row {data_start_row} to {max_row}")
            logging.info("Note: Conditional formatting disabled - differences not highlighted")
            
        except Exception as e:
            logging.error(f"Error during sheet comparison check: {e}")

    def _apply_span_length_tolerance(self, excel_span, qc_span, tolerance):
        """
        Check if QC span length is within tolerance of Excel span length
        
        Args:
            excel_span (str): Span length from Excel data
            qc_span (str): Span length from QC data
            tolerance (float): Tolerance in feet
            
        Returns:
            str: QC span length if within tolerance, otherwise Excel span length
        """
        if not qc_span or not excel_span:
            return excel_span or qc_span
        
        try:
            # Convert both to numeric values - remove commas, single quotes, and extra spaces
            excel_clean = str(excel_span).replace(',', '').replace("'", '').strip()
            qc_clean = str(qc_span).replace(',', '').replace("'", '').strip()
            
            excel_value = float(excel_clean)
            qc_value = float(qc_clean)
            
            # Check if difference is within tolerance
            difference = abs(excel_value - qc_value)
            if difference <= tolerance:
                logging.info(f"Using QC span length {qc_span} (Excel: {excel_span}, difference: {difference:.1f}ft, tolerance: {tolerance}ft)")
                return qc_span
            else:
                logging.info(f"QC span length {qc_span} outside tolerance (Excel: {excel_span}, difference: {difference:.1f}ft, tolerance: {tolerance}ft) - using Excel value")
                return excel_span
                
        except (ValueError, TypeError) as e:
            logging.debug(f"Error comparing span lengths '{excel_span}' vs '{qc_span}': {e}")
            return excel_span or qc_span