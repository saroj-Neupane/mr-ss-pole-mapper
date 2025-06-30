import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import shutil
import tempfile
import os
from .utils import Utils

class TensionCalculator:
    """Handles tension calculations using the Excel calculator file"""
    
    def __init__(self, calculator_file_path="Test Files/Metronet tension calculator.xlsm", worksheet_name="Calculations"):
        self.calculator_file_path = Path(calculator_file_path)
        self.worksheet_name = worksheet_name
        
        # Cell locations in the calculator
        self.cells = {
            'span_length': 'B2',
            'span_sag': 'E2', 
            'cable_installation': 'M4',
            'result_tension': 'R12'
        }
        
    def calculate_tension(self, span_length, attachment_height, midspan_height):
        """
        Calculate tension using the Excel calculator
        
        Args:
            span_length (float): Span length in feet
            attachment_height (float): Attachment height in feet
            midspan_height (float): Midspan height in feet
            
        Returns:
            float: Calculated tension value, or None if calculation fails
        """
        try:
            # Ensure span length is numeric
            span_length = float(str(span_length).replace("'", "").replace('"', "").strip())

            attachment_height_dec = Utils.parse_height_decimal(attachment_height)
            midspan_height_dec = Utils.parse_height_decimal(midspan_height)

            if attachment_height_dec is None or midspan_height_dec is None:
                raise ValueError("Unable to parse attachment or midspan height")

            attachment_height = round(attachment_height_dec, 2)
            midspan_height = round(midspan_height_dec, 2)
            
            # Calculate span sag
            span_sag = attachment_height - midspan_height
            cable_installation = attachment_height
            
            logging.info(f"Calculating tension with parsed values:")
            logging.info(f"  - Span Length: {span_length}' (raw: {span_length})")
            logging.info(f"  - Attachment Height: {attachment_height}' (raw: {attachment_height})")
            logging.info(f"  - Midspan Height: {midspan_height}' (raw: {midspan_height})")
            logging.info(f"  - Calculated Span Sag: {span_sag}' (attachment - midspan)")
            logging.info(f"  - Cable Installation: {cable_installation}' (same as attachment)")
            
            # Check if calculator file exists
            if not self.calculator_file_path.exists():
                logging.error(f"Tension calculator file not found: {self.calculator_file_path}")
                return None
            
            # Create a temporary copy of the calculator file to avoid modifying the original
            with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as temp_file:
                temp_path = Path(temp_file.name)
            
            try:
                # Copy the calculator file to temp location
                shutil.copy2(self.calculator_file_path, temp_path)
                
                # Open the temporary file
                wb = load_workbook(temp_path, keep_vba=True)
                
                if self.worksheet_name not in wb.sheetnames:
                    logging.error(f"Worksheet '{self.worksheet_name}' not found in calculator file")
                    return None
                
                ws = wb[self.worksheet_name]
                
                # Write input values to the specified cells
                ws[self.cells['span_length']] = span_length
                ws[self.cells['span_sag']] = span_sag
                ws[self.cells['cable_installation']] = cable_installation
                
                logging.info(f"Values written to Excel cells:")
                logging.info(f"  - {self.cells['span_length']} (Span Length) = {span_length}")
                logging.info(f"  - {self.cells['span_sag']} (Span Sag) = {span_sag}")
                logging.info(f"  - {self.cells['cable_installation']} (Cable Installation) = {cable_installation}")
                
                # Save the file to trigger any formulas
                wb.save(temp_path)
                wb.close()
                
                # Reopen the file to get calculated results
                wb = load_workbook(temp_path, data_only=True)
                ws = wb[self.worksheet_name]
                
                # Read the tension result
                tension_result = ws[self.cells['result_tension']].value
                
                wb.close()
                
                if tension_result is not None:
                    try:
                        tension_value = float(tension_result)
                        logging.info(f"Successfully calculated tension: {tension_value:.1f} lbs")
                        return tension_value
                    except (ValueError, TypeError):
                        logging.error(f"Invalid tension result: {tension_result}")
                        return None
                else:
                    logging.warning("Tension result cell is empty")
                    return None
                    
            finally:
                # Clean up temporary file
                try:
                    temp_path.unlink()
                except:
                    pass
                    
        except Exception as e:
            logging.error(f"Error calculating tension: {str(e)}")
            return None
    
    def calculate_tension_for_provider(self, provider_data, span_length):
        """
        Calculate tension for a specific provider given their attachment data
        
        Args:
            provider_data (dict): Dictionary containing attachment and midspan heights
            span_length (float): Span length in feet
            
        Returns:
            float: Calculated tension value, or None if calculation fails
        """
        try:
            # Extract attachment height and midspan height from provider data
            attachment_height = None
            midspan_height = None
            
            # Look for attachment height
            for key, value in provider_data.items():
                if 'attachment' in key.lower() or key.endswith('Attachment Ht'):
                    if value and str(value).strip():
                        attachment_height = self._parse_height_value(value)
                        break
            
            # Look for midspan height  
            for key, value in provider_data.items():
                if 'midspan' in key.lower() or key.endswith('Midspan Ht'):
                    if value and str(value).strip():
                        midspan_height = self._parse_height_value(value)
                        break
            
            if attachment_height is None:
                logging.warning("No attachment height found for tension calculation")
                return None
                
            if midspan_height is None:
                logging.warning("No midspan height found for tension calculation")
                return None
                
            if span_length is None or span_length <= 0:
                logging.warning("Invalid span length for tension calculation")
                return None
            
            return self.calculate_tension(span_length, attachment_height, midspan_height)
            
        except Exception as e:
            logging.error(f"Error calculating tension for provider: {str(e)}")
            return None
    
    def _parse_height_value(self, value):
        """Parse height value from various formats"""
        try:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return None

            decimal_feet = Utils.parse_height_decimal(value)

            if decimal_feet is not None:
                return round(decimal_feet, 2)

            # Fallback: strip quotes/units and try float conversion
            value_str = str(value).replace("'", "").replace('"', "").strip()
            return float(value_str)

        except (ValueError, TypeError):
            logging.warning(f"Could not parse height value: {value}")
            return None
    
    def validate_calculator_file(self):
        """Validate that the calculator file exists and has the required structure"""
        try:
            if not self.calculator_file_path.exists():
                return False, f"Calculator file not found: {self.calculator_file_path}"
            
            wb = load_workbook(self.calculator_file_path, data_only=True)
            
            if self.worksheet_name not in wb.sheetnames:
                return False, f"Worksheet '{self.worksheet_name}' not found"
            
            ws = wb[self.worksheet_name]
            
            # Check if required cells exist (basic validation)
            for cell_name, cell_ref in self.cells.items():
                try:
                    cell = ws[cell_ref]
                except:
                    return False, f"Cell {cell_ref} ({cell_name}) not accessible"
            
            wb.close()
            return True, "Calculator file validated successfully"
            
        except Exception as e:
            return False, f"Error validating calculator file: {str(e)}"
